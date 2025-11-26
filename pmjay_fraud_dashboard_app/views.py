from .models import Last24Hour, SuspiciousHospital, HospitalBeds, UploadHistory
from openpyxl.styles import PatternFill, Font, Border, Side
from django.contrib.auth.decorators import login_required
from django.db.models import Case, When, Value, CharField
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.http import require_http_methods
from django.template.loader import render_to_string
from django.db.models import Sum, Q, F, Func, Count, Subquery, OuterRef, Value, BooleanField, Exists, IntegerField
import random
from django.http import JsonResponse, HttpResponse
from django.utils.timezone import timedelta
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
import datetime
from django.db.models.functions import Cast
pd = None
np = None

from .utils.date_helpers import parse_date
from .utils.constants import SHAPEFILE_DISTRICT_MAPPING
from .features.dashboard.selectors import get_watchlist_base_query as patient_admitted_in_watchlist_hospital_base_query
from .utils.pandas_loader import load_dataframes
from .utils.logging import Timer

class HTML:
    def __init__(self, string=None, base_url=None): pass
    def write_pdf(self, target=None): pass

from django.shortcuts import render
from django.db.models.functions import TruncDate
from django.utils import timezone
import io
from django.contrib import messages
from django.db import transaction, connections
import sys
from collections import defaultdict
import time
import logging

logger = logging.getLogger(__name__)

class Timer:
    def __init__(self, label):
        self.label = label

    def __enter__(self):
        self.start = time.perf_counter()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        elapsed = time.perf_counter() - self.start
        logger.info(f"[TIMER-LOG] {self.label}: {elapsed:.2f}s")



# Helper to close DB connections and reduce lock contention
def close_db_connection():
    """Close the database connection after request completes"""
    connections.close_all()

def login_view(request):
    # If they’re already logged in, send them straight to dashboard
    if request.user.is_authenticated:
        return redirect('dashboard')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        print(f"LOGIN ATTEMPT: username='{username}', password length={len(password) if password else 0}")
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user is not None:
            login(request, user)
            # call_command('process_new_files')
            return redirect('dashboard')
        else:
            error = "Invalid username or password"

    return render(request, 'login.html', {'error': error})

def logout_view(request):
    logout(request)
    return redirect('login')

class Upper(Func):
    function = 'UPPER'
    template = '%(function)s(%(expressions)s)'

@login_required
def import_data_view(request):
    if request.method == 'POST':
        with Timer("import_data_view TOTAL"):
            uploaded_files = request.FILES.getlist('files')
            print(f"\n{'='*80}\nIMPORT DATA STARTED\n{'='*80}\n")
            
            try:
                new_records = 0
                for uploaded_file in uploaded_files:
                    try:
                        with Timer(f"Read file {uploaded_file.name}"):
                            # Detect file type and read with proper datetime handling
                            if uploaded_file.name.lower().endswith('.csv'):
                                df = pd.read_csv(uploaded_file, dtype=str)
                            else:
                                # For Excel files: Read WITHOUT dtype or parse_dates to preserve native types
                                # Excel stores datetimes as native datetime objects; pandas will auto-detect them
                                # We'll normalize column names AFTER reading so parse_dates matching isn't an issue
                                df = pd.read_excel(uploaded_file)
                        
                        print(f"\n>>> Processing file: {uploaded_file.name}")
                        print(f">>> Total rows in file: {len(df)}")

                        with Timer("Normalize & validate columns: Import Data"):
                            # Normalize column names
                            df.columns = df.columns.str.strip().str.lower()
                            
                            # Map CSV columns to model fields (case-insensitive)
                            column_mapping = {
                                'hospital_code': 'hospital_id',  # ← KEY FIX: CSV uses 'hospital_code', model uses 'hospital_id'
                                'hosp_state_name': 'hosp_state_name',
                            }
                            
                            # Rename columns according to mapping
                            df_renamed = df.rename(columns=column_mapping)
                            
                            # Expected model fields
                            required_columns = [
                                'registration_id', 
                                'admission_dt',
                                'preauth_init_date',
                                'hospital_id',  # Note: this comes from 'hospital_code' in CSV
                                'amount_claim_initiated',
                                'hospital_type',
                                'case_type',
                                'patient_state_name',
                                'age',
                                'procedure_code',
                                'category_details',
                                'patient_district_name',
                                'patient_name',
                                'gender',
                                'hospital_name',
                                'hosp_state_name',
                                'family_id'
                            ]
                            
                            missing_cols = [col for col in required_columns if col not in df_renamed.columns]
                            if missing_cols:
                                print(f"❌ SKIPPED: Missing columns: {missing_cols}")
                                messages.error(request, f"Skipped {uploaded_file.name}: Missing columns {', '.join(missing_cols)}")
                                continue
                        
                        # List of all datetime columns to process
                        datetime_columns = ['preauth_init_date', 'admission_dt', 'discharge_dt', 'surgery_dt', 
                                        'preauth_approved_date', 'preauth_rejected_date', 'payment_paid_dt', 
                                        'death_dt', 'claim_init_date', 'cpd_approved_date', 'aco_approved_date', 
                                        'sha_approved_date', 'claim_approved_date', 'last_insert_dt']
                        
                        with Timer("Datetime parsing: Import Data"):
                            for col in datetime_columns:
                                if col in df_renamed.columns:
                                    # If already datetime (from Excel's native format), keep as-is
                                    if pd.api.types.is_datetime64_any_dtype(df_renamed[col]):
                                        print(f"  ✓ {col} already datetime format (native from Excel)")
                                    else:
                                        # If numeric or string, try to parse as datetime
                                        # infer_datetime_format=True helps with various date formats
                                        print(f"  → Converting {col} from {df_renamed[col].dtype} to datetime...")
                                        df_renamed[col] = pd.to_datetime(df_renamed[col], errors='coerce', infer_datetime_format=True)

                        with Timer("Import Data: Drop invalid rows"):
                            missing_registration = df_renamed[df_renamed['registration_id'].isna() | (df_renamed['registration_id'].astype(str).str.strip() == '')]
                            missing_preauth = df_renamed[df_renamed['preauth_init_date'].isna()]
                            
                            print(f"    - Rows with missing/empty registration_id: {len(missing_registration)}")
                            print(f"    - Rows with missing/invalid preauth_init_date: {len(missing_preauth)}")
                            
                            # Show sample of problematic rows
                            if len(missing_preauth) > 0 and len(missing_preauth) <= 10:
                                print(f"\n      Sample of rows with invalid preauth_init_date:")
                                for idx, row in missing_preauth.head(5).iterrows():
                                    print(f"       Row {idx}: registration_id={row.get('registration_id')}, preauth_init_date_raw={row.get('preauth_init_date')}")
                            elif len(missing_preauth) > 10:
                                print(f"\n      First 5 rows with invalid preauth_init_date:")
                                for idx, row in missing_preauth.head(5).iterrows():
                                    print(f"       Row {idx}: registration_id={row.get('registration_id')}")
                            
                            # Drop rows with missing critical fields
                            valid_df = df_renamed.dropna(subset=['registration_id', 'preauth_init_date'])
                        dropped_count = len(df_renamed) - len(valid_df)
                        print(f"\n>>> Rows ready for import: {len(valid_df)} (dropped {dropped_count} incomplete rows)")
                        
                        if len(valid_df) == 0:
                            print("No valid rows to import after cleaning")
                            messages.error(request, f"No valid records in {uploaded_file.name}")
                            continue

                        # Prepare model instances
                        with Timer("Import Data: Row to model instance creation"):
                            instances = []
                            for idx, row in valid_df.iterrows():
                                try:
                                    admission_dt = row.get('admission_dt')
                                    if pd.isna(admission_dt):
                                        admission_dt = None
                                    elif isinstance(admission_dt, pd.Timestamp):
                                        admission_dt = admission_dt.to_pydatetime()
                                    
                                    instances.append(Last24Hour(
                                        registration_id=row.get('registration_id'),
                                        case_id=row.get('case_id'),
                                        member_id=row.get('member_id'),
                                        family_id=row.get('family_id'),
                                        patient_name=row.get('patient_name'),
                                        patient_dob=row.get('patient_dob'),
                                        patient_state_code=row.get('patient_state_code'),
                                        patient_district_code=row.get('patient_district_code'),
                                        patient_district_name=row.get('patient_district_name'),
                                        patient_state_name=row.get('patient_state_name'),
                                        gender=row.get('gender'),
                                        age=int(row.get('age')) if pd.notnull(row.get('age')) else None,
                                        policy_code=row.get('policy_code'),
                                        renewal_code=row.get('renewal_code'),
                                        category_details=row.get('category_details'),
                                        speciality_code=row.get('speciality_code'),
                                        procedure_details=row.get('procedure_details'),
                                        procedure_code=row.get('procedure_code'),
                                        case_type=row.get('case_type'),
                                        status_id_pk=row.get('status_id_pk'),
                                        case_status=row.get('case_status'),
                                        hospital_id=row.get('hospital_id'),
                                        hospital_name=row.get('hospital_name'),
                                        hosp_district_name=row.get('hosp_district_name'),
                                        hosp_state_name=row.get('hosp_state_name'),
                                        hospital_state_cd=row.get('hospital_state_cd'),
                                        hospital_district_cd=row.get('hospital_district_cd'),
                                        hosp_pan_number=row.get('hosp_pan_number'),
                                        hospital_type=row.get('hospital_type'),
                                        admission_dt=admission_dt,
                                        preauth_init_date=row.get('preauth_init_date'),
                                        amount_preauth_initiated=float(row.get('amount_preauth_initiated')) if pd.notnull(row.get('amount_preauth_initiated')) else None,
                                        preauth_approved_date=row.get('preauth_approved_date'),
                                        amount_preauth_approved=float(row.get('amount_preauth_approved')) if pd.notnull(row.get('amount_preauth_approved')) else None,
                                        preauth_rejected_date=row.get('preauth_rejected_date'),
                                        surgery_dt=row.get('surgery_dt'),
                                        death_dt=row.get('death_dt'),
                                        discharge_dt=row.get('discharge_dt'),
                                        claim_init_date=row.get('claim_init_date'),
                                        amount_claim_initiated=float(row.get('amount_claim_initiated')) if pd.notnull(row.get('amount_claim_initiated')) else None,
                                        amount_claim_approved=float(row.get('amount_claim_approved')) if pd.notnull(row.get('amount_claim_approved')) else None,
                                        amount_claim_paid=float(row.get('amount_claim_paid')) if pd.notnull(row.get('amount_claim_paid')) else None,
                                        claim_rejected_date=row.get('claim_rejected_date'),
                                        rf_amount=float(row.get('rf_amount')) if pd.notnull(row.get('rf_amount')) else None,
                                        tds_amount=float(row.get('tds_amount')) if pd.notnull(row.get('tds_amount')) else None,
                                        utr_no=row.get('utr_no'),
                                        payment_paid_dt=row.get('payment_paid_dt'),
                                        transaction_amount=float(row.get('transaction_amount')) if pd.notnull(row.get('transaction_amount')) else None,
                                        cpd_approved_date=row.get('cpd_approved_date'),
                                        cpd_approved_amount=float(row.get('cpd_approved_amount')) if pd.notnull(row.get('cpd_approved_amount')) else None,
                                        cpd_user=row.get('cpd_user'),
                                        aco_approved_date=row.get('aco_approved_date'),
                                        aco_approved_amount=float(row.get('aco_approved_amount')) if pd.notnull(row.get('aco_approved_amount')) else None,
                                        aco_user=row.get('aco_user'),
                                        sha_approved_date=row.get('sha_approved_date'),
                                        sha_approved_amount=float(row.get('sha_approved_amount')) if pd.notnull(row.get('sha_approved_amount')) else None,
                                        sha_user=row.get('sha_user'),
                                        json_object_perauth=row.get('json_object_perauth'),
                                        json_object_claim=row.get('json_object_claim'),
                                        json_object_ben=row.get('json_object_ben'),
                                        src_account_no=row.get('src_account_no'),
                                        src_ifsc_code=row.get('src_ifsc_code'),
                                        paid_flag=row.get('paid_flag'),
                                        hosp_account_number=row.get('hosp_account_number'),
                                        ben_ifsc_code=row.get('ben_ifsc_code'),
                                        current_workflow_role=row.get('current_workflow_role'),
                                        current_workflow_user=row.get('current_workflow_user'),
                                        service_request_type=row.get('service_request_type'),
                                        m_flag=row.get('m_flag'),
                                        careplan_desc=row.get('careplan_desc'),
                                        discharge_type=row.get('discharge_type'),
                                        admission_type=row.get('admission_type'),
                                        claim_approved_date=row.get('claim_approved_date'),
                                        last_insert_dt=row.get('last_insert_dt'),
                                    ))
                                except Exception as row_err:
                                    print(f"❌ Error creating instance for row {idx}: {row_err}")
                                    continue

                        print(f">>> Created {len(instances)} model instances")
                        
                        if len(instances) == 0:
                            print("❌ No instances created")
                            continue
                        
                        # Bulk create with error handling
                        try:
                            with Timer("Import Data: bulk_create Last24Hour"):
                                result = Last24Hour.objects.bulk_create(
                                    instances,
                                    batch_size=500,
                                    ignore_conflicts=True
                                )
                                new_records += len(result)
                                print(f"✅ Successfully inserted {len(result)} records")
                        except Exception as bulk_err:
                            print(f"❌ Bulk create error: {bulk_err}")
                            messages.error(request, f"Database error: {str(bulk_err)}")
                            continue
                    
                    except Exception as file_err:
                        print(f"❌ Error processing {uploaded_file.name}: {file_err}")
                        messages.error(request, f"Error processing {uploaded_file.name}: {str(file_err)}")
                        import traceback
                        traceback.print_exc()
                        continue

                print(f"\n{'='*80}")
                print(f"IMPORT COMPLETED: {new_records} total records imported")
                print(f"{'='*80}\n")
                messages.success(request, f"✅ Imported {new_records} new records successfully!")
            except Exception as e:
                print(f"❌ System error: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"System error: {str(e)}")

        return redirect('dashboard')
    
    return redirect('dashboard')
def data_management(request):
    with Timer("data_management TOTAL"):
        with Timer("UploadHistory.in_bulk"):
            histories = UploadHistory.objects.in_bulk(field_name='model_type')
        return render(request, 'data_management.html', {
            'active_page': 'data_management',
            'file_upload_histories': histories,
        })

@require_POST
def upload_management_data(request):
    with Timer("upload_management_data TOTAL"):
        file = request.FILES.get('file')
        model_type = request.POST.get('model_type')
        
        if not file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)
        if not model_type:
            return JsonResponse({'status': 'error', 'message': 'Missing model type'}, status=400)

        try:
            with Timer("Read Excel file"):
                df = pd.read_excel(file)

            with Timer("Replace NaN"):
                df = df.replace({np.nan: None})
            
            with Timer("Validate required columns"):
                required_columns = {
                    'suspicious': ['hospital_id', 'hospital_name', 'number_of_surgeons', 'number_of_ot'],
                    'beds': ['hospital_id', 'bed_strength', 'hospital_district', 'hospital_email_id', 'hospital_name', 'number_of_surgeons', 'number_of_ot']
                }.get(model_type, [])

                if not required_columns:
                    return JsonResponse({'status': 'error', 'message': 'Invalid model type'}, status=400)

                missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Missing columns: {", ".join(missing_cols)}'
                }, status=400)

            # Handle duplicates before processing
            if model_type == 'suspicious':
                with Timer("Drop duplicates (suspicious)"):
                    # Remove duplicates keeping first occurrence
                    initial_count = len(df)
                    df = df.drop_duplicates(
                        subset=['hospital_id'], 
                        keep='first' 
                    )
                removed_duplicates = initial_count - len(df)
                
                with Timer("DB write SuspiciousHospital"):
                    with transaction.atomic():
                        SuspiciousHospital.objects.all().delete()
                        with Timer("Build SuspiciousHospital instances"):
                            hospitals = [
                                SuspiciousHospital(
                                    hospital_id=row['hospital_id'],
                                    hospital_name=row['hospital_name'],
                                    number_of_surgeons=row['number_of_surgeons'],
                                    number_of_ot=row['number_of_ot']
                                )
                                for _, row in df.iterrows()
                            ]
                        SuspiciousHospital.objects.bulk_create(hospitals)
                        UploadHistory.objects.update_or_create(
                            model_type='suspicious',
                            defaults={'filename': file.name}
                        )
                        
                        return JsonResponse({
                            'status': 'success',
                            'message': f'Uploaded {len(df)} records (removed {removed_duplicates} duplicates)'
                        })

            elif model_type == 'beds':
                # Handle Hospital Beds upload
                required_columns = ['hospital_id', 'hospital_district', 'hospital_email_id', 'bed_strength', 'number_of_surgeons', 'number_of_ot']
                with Timer("Convert bed_strength to int"):
                    # Fill missing bed_strength with 0
                    df['bed_strength'] = df['bed_strength'].fillna(0)
                    
                    # Convert to integers with error handling
                    try:
                        df['bed_strength'] = df['bed_strength'].astype(int)
                    except ValueError:
                        return JsonResponse({
                            'status': 'error',
                            'message': 'bed_strength contains non-numeric values that cannot be converted to integers'
                        }, status=400)

                with Timer("Drop duplicates (beds)"):
                    # Remove duplicates
                    initial_count = len(df)
                    df = df.drop_duplicates(subset=['hospital_id'], keep='first')
                    removed_duplicates = initial_count - len(df)

                with Timer("DB write HospitalBeds"):
                    with transaction.atomic():
                        HospitalBeds.objects.all().delete()
                        with Timer("Build HospitalBeds instances"):
                            beds = []
                            for excel_row_no, (_idx, row) in enumerate(df.iterrows(), start=2):
                                try:
                                    beds.append(HospitalBeds(
                                        hospital_id=row['hospital_id'],
                                        hospital_name=row.get('hospital_name', ''),
                                        hospital_district=row.get('hospital_district'),
                                        hospital_email_id=row.get('hospital_email_id'),
                                        bed_strength=row['bed_strength'],
                                        number_of_surgeons=row['number_of_surgeons'],
                                        number_of_ot=row['number_of_ot']
                                    ))
                                except Exception as e:
                                    return JsonResponse({
                                        'status': 'error',
                                        'message': f'Error in row {+2}: {str(e)}'
                                    }, status=400)
                                    
                            HospitalBeds.objects.bulk_create(beds)
                            UploadHistory.objects.update_or_create(
                                model_type='beds',
                                defaults={'filename': file.name}
                            )
                    
                return JsonResponse({
                    'status': 'success',
                    'message': f'Uploaded {len(df)} beds records. {removed_duplicates} duplicates removed. {initial_count - len(df)} nulls converted to 0.'
                })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'{str(e)} (Line {sys.exc_info()[-1].tb_lineno if hasattr(sys, "exc_info") else "N/A"})'
            }, status=400)
    
@login_required
def latest_uploads(request):
    """
    Return JSON mapping of model_type -> {filename, uploaded_at}
    """
    with Timer("latest_uploads TOTAL"):
        with Timer("UploadHistory.objects.all"):
            qs = UploadHistory.objects.all()
    with Timer("Build JSON payload"):
        data = {}
        for hist in qs:
            data[hist.model_type] = {
                'filename': hist.filename,
                # ISO timestamp is easiest for JS to parse, but we can shorten if you like
                'uploaded_at': hist.uploaded_at.strftime('%Y-%m-%d %H:%M'),
            }
    return JsonResponse(data)

def get_management_data(request):
    with Timer("get_management_data TOTAL"):
        model_type = request.GET.get('model_type')
    
    try:
        with Timer("Fetch management data"):
            if model_type == 'suspicious':
                data = SuspiciousHospital.objects.all()
                fields = ['district', 'hospital_id', 'hospital_name', 'number_of_surgeons', 'number_of_ot']
                labels = ['district', 'hospital_id', 'hospital_name', 'number_of_surgeons', 'number_of_ot']
            elif model_type == 'beds':
                data = HospitalBeds.objects.all()
                fields = ['hospital_id', 'hospital_name', 'hospital_district', 'hospital_email_id', 'bed_strength', 'number_of_surgeons', 'number_of_ot']
                labels = ['hospital_id', 'hospital_name', 'hospital_district', 'hospital_email_id', 'bed_strength', 'number_of_surgeons', 'number_of_ot']
            else:
                return HttpResponse('Invalid model type')
            
        # build the table head
        with Timer("Build HTML table"):
            html = '<table><thead><tr>'
            for label in labels:
                html += f'<th>{label}</th>'
            html += '</tr></thead><tbody>'

            # build the rows
            for item in data:
                html += '<tr>'
                for field in fields:
                    html += f'<td>{getattr(item, field)}</td>'
                html += '</tr>'
            html += '</tbody></table>'
    except Exception as e:
        return HttpResponse(f'Error: {str(e)}', status=400)
    
    return HttpResponse(html)



def get_ot_overflow_hospital_ids(start_date, end_date, districts=None):
    with Timer("get_ot_overflow_hospital_ids TOTAL"):

        with Timer("get_ot_overflow_hospital_ids: Load capacity map"):
            cap_map = {
                rec['hospital_id']: rec['number_of_surgeons'] * 30
                for rec in HospitalBeds.objects.filter(number_of_surgeons__isnull=False).values('hospital_id', 'number_of_surgeons')
            }
        with Timer("get_ot_overflow_hospital_ids: Fetch relevant cases"):
            qs = Last24Hour.objects.filter(
                hospital_type='P',
                category_details__contains='Opthalmology',
                preauth_init_date__date__gte=start_date,
                preauth_init_date__date__lte=end_date
            )
            # print(qs)
            if districts:
                qs = qs.filter(patient_district_name__in=districts)
            # Fetch all relevant cases at once
            all_cases = list(qs.values('id', 'hospital_id', 'preauth_init_date'))   

        with Timer("get_ot_overflow_hospital_ids: Group cases by hospital"):
            cases_by_hospital = defaultdict(list)
            for case in all_cases:
                cases_by_hospital[case['hospital_id']].append(case)
        
        with Timer("get_ot_overflow_hospital_ids: Detect OT overflow"):
            flagged_ot_ids = set()
            for hid, cap in cap_map.items():
                cases = sorted(
                    cases_by_hospital.get(hid, []),
                    key=lambda x: x['preauth_init_date']
                )
                if len(cases) > cap:
                    flagged_ot_ids.update([c['id'] for c in cases[cap:]])
    # print(len(flagged_ot_ids))
    return flagged_ot_ids





# Add to views.py





def download_flagged_claims_excel(request):
    with Timer("download_flagged_claims_excel TOTAL"):

        with Timer("download_flagged_claims_excel: Parse params"):
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
            
            # 1. Apply same filters as other endpoints
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []

        with Timer("download_flagged_claims_excel: Build base queryset"):
            # 2. Build queryset with today's filter
            qs, suspicious_hospitals = patient_admitted_in_watchlist_hospital_base_query(start_date, end_date, districts)

        with Timer("download_flagged_claims_excel: Fetch & serialize rows"):
            # 3. Prepare data with required fields only
            rows = [{
                'Claim ID': case.registration_id or case.case_id,
                'Patient Name': case.patient_name or f"Patient {case.member_id}",
                'District': case.patient_district_name,
                'Preauth Initiated Date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
                'Preauth Initiated Time': case.preauth_init_date.strftime('%H:%M:%S') if case.preauth_init_date else 'N/A',
                'Hospital ID': case.hospital_id,
                'Hospital Name': case.hospital_name,
                'Amount': float(case.amount_claim_initiated) if case.amount_claim_initiated else 0.0,
                'Reason': 'Suspicious hospital',
                'Date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
            } for case in qs.only(
                'registration_id', 'case_id', 'patient_name', 'member_id', 'patient_district_name', 'preauth_init_date', 'preauth_init_date', 'hospital_id', 'hospital_name', 'amount_claim_initiated', 'preauth_init_date'
            )]

        # 4. Create DataFrame with defined column order
        columns = ['Claim ID', 'Patient Name', 'District', 'Preauth Initiated Date', 'Preauth Initiated Time', 'Hospital ID', 'Hospital Name',  
                'Amount', 'Reason', 'Date']
            
        with Timer("download_flagged_claims_excel: Create DataFrame"):
            df = pd.DataFrame(rows, columns=columns)

        with Timer("download_flagged_claims_excel: Write Excel file"):
            # 5. Efficient Excel styling
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Flagged Claims')
                workbook = writer.book
                worksheet = writer.sheets['Flagged Claims']
                
        # Style definitions
        red_fill = PatternFill(start_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF')
        
        with Timer("download_flagged_claims_excel: Apply Excel styling"):
            # Apply styling to entire Reason column
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=9, max_col=9):
                for cell in row:
                    cell.fill = red_fill
                    cell.font = white_font              

    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="flagged_claims_{start_date}_to_{end_date}.xlsx"'
    return response

@require_http_methods(["GET", "POST"])
def download_flagged_claims_report(request):
    with Timer("download_flagged_claims_report TOTAL"):
        with Timer("download_flagged_claims_report: Parse Params"):
            startDate = request.POST.get('start_date')
            endDate = request.POST.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)

            # 1) Read parameters & chart images
            district = request.POST.get('district', '')
            districts = district.split(',') if district else []

        with Timer("download_flagged_claims_report: Parse base64 images"):
            # Each value is "data:image/png;base64,XXXXX"
            def strip_prefix(data_url):
                return data_url.split('base64,', 1)[1]

            flagged_b64 = strip_prefix(request.POST.get('flagged_chart', ''))
            age_b64     = strip_prefix(request.POST.get('age_chart', ''))
            gender_b64  = strip_prefix(request.POST.get('gender_chart', ''))
            age_callouts    = request.POST.get('age_callouts', '')
            gender_callouts = request.POST.get('gender_callouts', '')
            map_b64 = strip_prefix(request.POST.get('flagged_map', ''))

        with Timer("download_flagged_claims_report: Build base queryset"):
            # 2) Fetch the FULL flagged-claims data (no pagination)
            qs, suspicious_hospitals = patient_admitted_in_watchlist_hospital_base_query(start_date, end_date, districts)

        with Timer("download_flagged_claims_report: Serialize table rows"):
            table_rows = []
            for idx, case in enumerate(qs, start=1):
                table_rows.append({
                    'serial_no':     idx,
                    'claim_id':      case.registration_id or case.case_id or 'N/A',
                    'patient_name':  case.patient_name or f"Patient {case.member_id}",
                    'patient_district_name': case.patient_district_name or 'N/A',
                    'preauth_initiated_date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
                    'preauth_initiated_time': case.preauth_init_date.strftime('%H:%M:%S') if case.preauth_init_date else 'N/A',
                    'hospital_id': case.hospital_id or 'N/A',
                    'hospital_name': case.hospital_name or 'N/A',
                    'amount':        case.amount_claim_initiated or 0,
                    'reason':        'Suspicious hospital'
                })
        report_districts = sorted({
            row['patient_district_name'] 
            for row in table_rows 
            if row['patient_district_name'] and row['patient_district_name'] != 'N/A'
        })

        # 3) Render HTML via a dedicated template
        context = {
            'logo_url':    request.build_absolute_uri('/static/images/pmjaylogo.png'),
            'title':       'SAFU DASHBOARD ANALYSIS REPORT',
            'date':        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'table_rows':  table_rows,
            'report_districts': report_districts,
            'flagged_b64': flagged_b64,
            'age_b64':     age_b64,
            'gender_b64':  gender_b64,
            'age_callouts':    age_callouts,
            'gender_callouts': gender_callouts,
            'map_b64': map_b64
        }

        with Timer("download_flagged_claims_report: Render HTML Template"):
            html_string = render_to_string('flagged_claims_report.html', context)

        with Timer("download_flagged_claims_report: Generate PDF"):
            # 4) Generate PDF
            html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
            pdf  = html.write_pdf()

    # 5) Return as attachment
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="flagged_claims_report.pdf"'
    return response

from pmjay_fraud_dashboard_app.features.high_value_claims.selectors import high_value_claims_base_query

def hospital_bed_cases_base_query():
    with Timer("hospital_bed_cases_base_query"):
        return dict(
            HospitalBeds.objects.values_list('hospital_id', 'bed_strength')
        )

def get_hospital_bed_cases(request):
    with Timer("get_hospital_bed_cases TOTAL"):
        with Timer("get_hospital_bed_cases: Load bed strengths"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
        with Timer("get_hospital_bed_cases: Load bed strengths"): 
            # 1. Load bed strengths 
            bed_strengths = hospital_bed_cases_base_query() 
            hospital_ids = list(bed_strengths.keys()) 
        with Timer("get_hospital_bed_cases: Build base queryset"): 
            # 2. Build base queryset: only private hospitals, filter by district if needed
            qs = ( Last24Hour.objects .filter( hospital_type='P', hospital_id__in=hospital_ids, admission_dt__date__gte=start_date, admission_dt__date__lte=end_date, ) ) 
            if districts: qs = qs.filter(patient_district_name__in=districts) 
        with Timer("get_hospital_bed_cases: Aggregate admissions per hospital/day"): 
            # 3. Aggregate admissions per hospital per day 
            admissions = ( qs .annotate(day=TruncDate('admission_dt')) .values('hospital_id', 'hospital_name', 'patient_district_name', 'hosp_state_name', 'day') .annotate(admissions=Count('id')) ) 
        with Timer("get_hospital_bed_cases: Build violation list"):
             #4. Flag violations: hospital exceeds bed strength on that day 
             violations = [ { 'hospital_id': adm['hospital_id'], 'hospital_name': adm['hospital_name'], 'district': adm['patient_district_name'], 'state': adm['hosp_state_name'], 'day': adm['day'].strftime('%Y-%m-%d') if adm['day'] else 'N/A', 'admissions': adm['admissions'], 'bed_capacity': bed_strengths.get(adm['hospital_id'], 0), 'excess': adm['admissions'] - bed_strengths.get(adm['hospital_id'], 0) } for adm in admissions if adm['admissions'] > bed_strengths.get(adm['hospital_id'], 0) ]
        with Timer("get_hospital_bed_cases: Compute summary counts"): 
            unique_hospitals = len({v['hospital_id'] for v in violations})
            # 5. Prepare summary counts 
            total_violations = len(violations) 
            yesterday = end_date - timedelta(days=1) 
            yesterday_violations = sum(1 for v in violations if v['day'] == yesterday.strftime('%Y-%m-%d')) 
            thirty_days_ago = end_date - timedelta(days=30) 
            last_30_days_violations = sum(1 for v in violations if v['day'] >= thirty_days_ago.strftime('%Y-%m-%d')) 
        return JsonResponse({ 
            'total': total_violations, 
            'unique_hospitals': unique_hospitals,
            'yesterday': yesterday_violations, 
            'last_30_days': last_30_days_violations, 
            'violations': violations })

def get_hospital_bed_details(request):
    with Timer("get_hospital_bed_details TOTAL"):

        with Timer("get_hospital_bed_details: Load bed strengths"):
            district_param = request.GET.get('district', '')
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 50))
            search_query = request.GET.get('search', '').strip()
            districts = district_param.split(',') if district_param else []

            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)

        with Timer("get_hospital_bed_details: Load Hospital Bed strength"):
            # 1. Load bed strengths
            bed_strengths = hospital_bed_cases_base_query()
            hospital_ids = list(bed_strengths.keys())
        with Timer("get_hospital_bed_details: Build base filtered queryset"):
            # 2. Aggregate admissions per hospital per day
            qs = (
                Last24Hour.objects
                .filter(
                    hospital_type='P',
                    hospital_id__in=hospital_ids,
                    admission_dt__date__gte=start_date,
                    admission_dt__date__lte=end_date,
                )
            )

            if districts:
                qs = qs.filter(patient_district_name__in=districts)
        with Timer("get_hospital_bed_details: Aggregate admissions"):
            admissions = (
                qs
                .annotate(day=TruncDate('admission_dt'))
                .values('hospital_id', 'hospital_name', 'patient_district_name', 'hosp_state_name', 'day')
                .annotate(admissions=Count('id'))
            )

        with Timer("get_hospital_bed_details: Build violations list"):
            # 3. Flag violations: hospital exceeds bed strength on that day
            violations = [
                {
                    'hospital_id': adm['hospital_id'],
                    'hospital_name': adm['hospital_name'],
                    'district': adm['patient_district_name'],
                    'state': adm['hosp_state_name'],
                    'day': adm['day'].strftime('%Y-%m-%d') if adm['day'] else 'N/A',
                    'admissions': adm['admissions'],
                    'bed_capacity': bed_strengths.get(adm['hospital_id'], 0),
                    'excess': adm['admissions'] - bed_strengths.get(adm['hospital_id'], 0)
                }
                for adm in admissions
                if adm['admissions'] > bed_strengths.get(adm['hospital_id'], 0)
            ]
        with Timer("get_hospital_bed_details: Search filter"):
            if search_query:
                search_terms = [t.strip().lower() for t in search_query.split(',') if t.strip()]
                filtered_violations = []
                for violation in violations:
                    violation_values = [
                        str(violation['hospital_id']).lower(),
                        str(violation['hospital_name']).lower(),
                        str(violation['district']).lower(),
                        str(violation['state']).lower(),
                        str(violation['day']).lower(),
                        str(violation['admissions']).lower(),
                        str(violation['bed_capacity']).lower(),
                        str(violation['excess']).lower()
                    ]
                    # Each term must match exactly one of the values (not substring)
                    if all(any(term == value for value in violation_values) for term in search_terms):
                        filtered_violations.append(violation)
                violations = filtered_violations
        with Timer("get_hospital_bed_details: Pagination"):
            # 4. Paginate in-memory list
            paginator = Paginator(violations, page_size)
            page_obj = paginator.get_page(page)
        with Timer("get_hospital_bed_details: Serialize data"):
            # 5. Serialize data
            data = []
            for idx, violation in enumerate(page_obj.object_list, 1):
                data.append({
                    'serial_no': (page_obj.number - 1) * page_size + idx,
                    'hospital_id': violation['hospital_id'],
                    'hospital_name': violation['hospital_name'],
                    'district': violation['district'],
                    'state': violation['state'],
                    'bed_capacity': violation['bed_capacity'],
                    'admissions': violation['admissions'],
                    'excess': violation['excess'],
                    'last_violation': violation['day']
                })

    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })

def hospital_violations_by_district(request):
    with Timer("hospital_violations_by_district TOTAL"):

        with Timer("hospital_violations_by_district: Parse params & dates"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []
            search_query = request.GET.get('search', '').strip()
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)

        with Timer("hospital_violations_by_district: Load Bed Strength"):
            # 1. Load bed strengths
            bed_strengths = hospital_bed_cases_base_query()
            hospital_ids = list(bed_strengths.keys())

        with Timer("hospital_violations_by_district: Build base filtered queryset"):
            # 2. Aggregate admissions per hospital per day
            qs = (
                Last24Hour.objects
                .filter(
                    hospital_type='P',
                    hospital_id__in=hospital_ids,
                    admission_dt__date__gte=start_date,
                    admission_dt__date__lte=end_date,
                )
            )
            if districts:
                qs = qs.filter(patient_district_name__in=districts)

        with Timer("hospital_violations_by_district: Aggregate admissions"):
            admissions = (
                qs
                .annotate(day=TruncDate('admission_dt'))
                .values('hospital_id', 'patient_district_name', 'day')
                .annotate(admissions=Count('id'))
            )

        with Timer("hospital_violations_by_district: Violation filtering"):
            # 3. Flag violations: hospital exceeds bed strength on that day
            violations = [
                adm['patient_district_name']
                for adm in admissions
                if adm['admissions'] > bed_strengths.get(adm['hospital_id'], 0)
            ]
        with Timer("hospital_violations_by_district: Counter aggregation"):
            # 4. Count violations per district
            from collections import Counter
            district_counts = Counter(violations)

    return JsonResponse({
        'districts': list(district_counts.keys()),
        'counts': list(district_counts.values())
    })

def get_hospital_bed_violations_geo(request):
    with Timer("hospital_violations_by_district TOTAL"):

        with Timer("hospital_violations_by_district: Parse params & dates"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []

            # Parse dates
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
        with Timer("hospital_violations_by_district: Load bed strengths"):
            # Load bed strengths
            bed_strengths = hospital_bed_cases_base_query()
            hospital_ids = list(bed_strengths.keys())

        with Timer("hospital_violations_by_district: Build base queryset"):
            # Base queryset: only private hospitals, filter by district if needed
            qs = (
                Last24Hour.objects
                .filter(
                    hospital_type='P',
                    hospital_id__in=hospital_ids,
                    admission_dt__date__gte=start_date,
                    admission_dt__date__lte=end_date,
                )
            )
            if districts:
                qs = qs.filter(patient_district_name__in=districts)
        with Timer("hospital_violations_by_district: Aggregate admissions"):
            # Aggregate admissions per hospital per day
            admissions = (
                qs
                .annotate(day=TruncDate('admission_dt'))
                .values('hospital_id', 'patient_district_name', 'day')
                .annotate(admissions=Count('id'))
            )
        with Timer("hospital_violations_by_district: Violation filtering"):
            # Flag violations: hospital exceeds bed strength on that day
            violations = [
                adm['patient_district_name']
                for adm in admissions
                if adm['admissions'] > bed_strengths.get(adm['hospital_id'], 0)
            ]
        with Timer("hospital_violations_by_district: Counter aggregation"):
            # Count violations per district
            from collections import Counter
            district_counts = Counter(violations)

            # Map to FID for geo output
            result = []
            for district, count in district_counts.items():
                fid = SHAPEFILE_DISTRICT_MAPPING.get(district.lower())
                if fid is not None:
                    result.append({'fid': fid, 'count': count})

    return JsonResponse(result, safe=False)

def get_family_id_cases(request):
    with Timer("get_family_id_cases TOTAL"):

        # -----------------------------
        # Parse params
        # -----------------------------
        with Timer("get_family_id_cases: Parse params"):
            districts = request.GET.get('district', '')
            districts = [d.strip() for d in districts.split(',')] if districts else []

            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)

            yesterday = end_date - timedelta(days=1)
            thirty_days_ago = end_date - timedelta(days=30)

        # -----------------------------
        # Base queryset
        # -----------------------------
        with Timer("get_family_id_cases: Base queryset"):
            base_qs = Last24Hour.objects.filter(
                hospital_type='P',
                preauth_init_date__date__range=(start_date, end_date)
            )
            if districts:
                base_qs = base_qs.filter(patient_district_name__in=districts)

        # -----------------------------
        # Suspicious families (>1 claim per family per day)
        # -----------------------------
        with Timer("get_family_id_cases: Suspicious families aggregation"):
            suspicious = (
                base_qs
                .annotate(day=TruncDate('preauth_init_date'))
                .values('family_id', 'day')
                .annotate(count=Count('id'))
                .filter(count__gt=1)
            )

        # -----------------------------
        # Totals
        # -----------------------------
        with Timer("get_family_id_cases: Totals"):
            violating_family_ids = {r['family_id'] for r in suspicious if r['family_id']}

            yesterday_count = (
                suspicious
                .filter(day=yesterday)
                .values('family_id')
                .distinct()
                .count()
            )

            last_30_days = (
                Last24Hour.objects
                .filter(
                    hospital_type='P',
                    family_id__in=violating_family_ids,
                    preauth_init_date__date__range=(thirty_days_ago, end_date)
                )
                .values('family_id')
                .distinct()
                .count()
            )

        # -----------------------------
        # Fetch hospitals in ONE query (no N+1)
        # -----------------------------
        with Timer("get_family_id_cases: Fetch hospitals"):
            hospital_map = {}
            hospital_rows = (
                Last24Hour.objects
                .filter(
                    hospital_type='P',
                    family_id__in=violating_family_ids,
                    preauth_init_date__date__range=(start_date, end_date)
                )
                .annotate(day=TruncDate('preauth_init_date'))
                .values('family_id', 'day', 'hospital_name')
                .distinct()
            )

            for row in hospital_rows:
                key = (row['family_id'], row['day'])
                hospital_map.setdefault(key, set()).add(row['hospital_name'])

        # -----------------------------
        # Build response
        # -----------------------------
        with Timer("get_family_id_cases: Build response"):
            violations = []
            for r in suspicious:
                key = (r['family_id'], r['day'])
                violations.append({
                    'family_id': r['family_id'],
                    'date': r['day'].strftime('%Y-%m-%d'),
                    'count': r['count'],
                    'hospitals': sorted(hospital_map.get(key, []))
                })
            total_claims = sum(v['count'] for v in violations) 

    return JsonResponse({
        'total': len(violating_family_ids),
        'total_claims': total_claims,
        'yesterday': yesterday_count,
        'last_30_days': last_30_days,
        'violations': violations
    })

def get_family_id_cases_details(request):
    with Timer("get_family_id_cases_details TOTAL"):

        with Timer("get_family_id_cases_details: Subquery suspicious families"):
            district_param = request.GET.get('district', '')
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 50))
            districts = district_param.split(',') if district_param else []
            
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
        with Timer("get_family_id_cases_details: Subquery suspicious families"):
            # Subquery: Get family_ids with more than 1 cases 
            suspicious_families = Last24Hour.objects.annotate(
                day=TruncDate('preauth_init_date')
            ).filter(
                day__range=(start_date, end_date)
            ).values('family_id', 'day').annotate(
                count=Count('id')
            ).filter(
                count__gt=1
            ).values('family_id')
        with Timer("get_family_id_cases_details: Main cases query"):
            # Main query: Cases for those suspicious families today
            cases = Last24Hour.objects.filter(
                hospital_type='P',
                family_id__in=Subquery(suspicious_families),
                preauth_init_date__date__gte=start_date,
                preauth_init_date__date__lte=end_date
            ).order_by('family_id', 'preauth_init_date')
            
            if districts:
                cases = cases.filter(patient_district_name__in=districts)
        with Timer("get_family_id_cases_details: Pagination"):
            # Pagination
            paginator = Paginator(cases, page_size)
            page_obj = paginator.get_page(page)
        with Timer("get_family_id_cases_details: Serialization"):
            # Serialize data
            data = []
            for idx, case in enumerate(page_obj.object_list, 1):
                data.append({
                    'serial_no': (page_obj.number - 1) * page_size + idx,
                    'family_id': case.family_id,
                    'claim_id': case.registration_id or case.case_id or 'N/A',
                    'patient_name': case.patient_name or f"Patient {case.member_id}",
                    'district_name': case.patient_district_name or 'N/A',
                    'preauth_initiated_date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
                    'preauth_initiated_time': case.preauth_init_date.strftime('%H:%M:%S') if case.preauth_init_date else 'N/A',
                    'hospital_id': case.hospital_id or 'N/A',
                    'hospital_name': case.hospital_name or 'N/A',
                    'date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
                })
    
    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })

def get_family_violations_by_district(request):
    with Timer("get_family_violations_by_district TOTAL"):
        with Timer("get_family_violations_by_district: Parse Params"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []
            
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)

        with Timer("get_family_violations_by_district: Subquery suspicious families"):
            # Subquery: Get family_ids with more than 2 cases today
            suspicious_families = Last24Hour.objects.annotate(
                day=TruncDate('preauth_init_date')
            ).filter(
                day__range=(start_date, end_date)
            ).values('family_id', 'day').annotate(
                count=Count('id')
            ).filter(
                count__gt=1
            ).values('family_id')

        with Timer("get_family_violations_by_district: Count results - Main query"):
            # Main query: Count number of unique families per district
            result = Last24Hour.objects.filter(
                hospital_type='P',
                family_id__in=Subquery(suspicious_families),
                preauth_init_date__date__gte=start_date,
                preauth_init_date__date__lte=end_date
            )
            
            if districts:
                result = result.filter(patient_district_name__in=districts)
        with Timer("get_family_violations_by_district: aggregate by family ID"):
            result = result.values('patient_district_name').annotate(
                family_count=Count('family_id', distinct=True)
            ).order_by('-family_count')
    
    return JsonResponse({
        'districts': [item['patient_district_name'] for item in result],
        'counts': [item['family_count'] for item in result]
    })


def get_family_violations_demographics(request, type):
    with Timer("Get Family violations by demography"):
        with Timer("Get Family violations by demography: Parse Params"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []
            
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)

        with Timer("Get Family violations by demography: Subquery for suspicious hospitals"):
            # Subquery: Get family_ids with more than 2 cases today
            suspicious_families = Last24Hour.objects.annotate(
                day=TruncDate('preauth_init_date')
            ).filter(
                day__range=(start_date, end_date)
            ).values('family_id', 'day').annotate(
                count=Count('id')
            ).filter(
                count__gt=1
            ).values('family_id')
        with Timer("Get Family violations by demography: Base query"):
            # Base query for demographics
            base_query = Last24Hour.objects.filter(
                hospital_type='P',
                family_id__in=Subquery(suspicious_families),
                preauth_init_date__date__gte=start_date,
                preauth_init_date__date__lte=end_date
            )
        with Timer("Get Family violations by demography: Apply base filters"):
            if districts:
                base_query = base_query.filter(patient_district_name__in=districts)
            
            if type == 'age':
                with Timer("Get Family violations by demography: By age"):
                    age_groups = Case(
                        When(age__lt=20, then=Value('≤20')),
                        When(age__gte=20, age__lt=30, then=Value('21-30')),
                        When(age__gte=30, age__lt=40, then=Value('31-40')),
                        When(age__gte=40, age__lt=50, then=Value('41-50')),
                        When(age__gte=50, age__lt=60, then=Value('51-60')),
                        When(age__gte=60, then=Value('60+')),
                        default=Value('Unknown'),
                        output_field=CharField()
                    )
                    with Timer("Get Family violations by demography: Aggregate by age"):
                        age_data = base_query.annotate(age_group=age_groups).values('age_group') \
                            .annotate(count=Count('id')).order_by('age_group')
                
                        categories = ['≤20', '21-30', '31-40', '41-50', '51-60', '60+', 'Unknown']
                        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40', '#C9CBCF']
                
                        age_dict = {item['age_group']: item['count'] for item in age_data}
                        
                        data = {
                            'labels': categories,
                            'data': [age_dict.get(cat, 0) for cat in categories],
                            'colors': colors
                        }
            elif type == 'gender':
                with Timer("Get Family violations by demography: By Gender"):
                    gender_data = base_query.values('gender') \
                        .annotate(count=Count('id')).order_by('gender')
                    
                    categories = ['Male', 'Female', 'Other', 'Unknown']
                    colors = ['#36A2EB', '#FF6384', '#4BC0C0', '#C9CBCF']
                    
                    gender_map = {
                        'M': 'Male',
                        'F': 'Female',
                        'O': 'Other'
                    }
                    
                    gender_dict = {}
                    for item in gender_data:
                        gender = gender_map.get(item['gender'], 'Unknown')
                        gender_dict[gender] = item['count']
                    
                    data = {
                        'labels': categories,
                        'data': [gender_dict.get(cat, 0) for cat in categories],
                        'colors': colors
                    }
    
    return JsonResponse(data)

def get_family_violations_geo(request):
    with Timer("Get family violations by geo"):
        with Timer("Get family violations by geo: Parse Params"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []

            # parse dates
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
        with Timer("Get family violations by geo: Build base query"):
            # base queryset
            qs = Last24Hour.objects.filter(
                hospital_type='P',
                preauth_init_date__date__gte=start_date,
                preauth_init_date__date__lte=end_date
            )
        with Timer("Get family violations by geo: Build subquery for suspicious cases"):
            # Subquery: Get family_ids with more than 2 cases today
            suspicious_families = Last24Hour.objects.annotate(
                day=TruncDate('preauth_init_date')
            ).filter(
                day__range=(start_date, end_date)
            ).values('family_id', 'day').annotate(
                count=Count('id')
            ).filter(
                count__gt=1
            ).values('family_id')

            qs = qs.filter(family_id__in=Subquery(suspicious_families))

            if districts:
                qs = qs.filter(patient_district_name__in=districts)
        with Timer("Get family violations by geo: Aggreagte by Patient Name"):
            # aggregate by patient_district_name
            agg = qs.values('patient_district_name').annotate(count=Count('family_id', distinct=True))
        with Timer("Get family violations by geo: Map to FID"):
            # map to FID
            result = []
            for row in agg:
                fid = SHAPEFILE_DISTRICT_MAPPING.get(row['patient_district_name'].lower())
                if fid is not None:
                    result.append({'fid': fid, 'count': row['count']})

    return JsonResponse(result, safe=False)

def get_geo_anomalies(request):
    with Timer("Get Geo Anomalies Case Type"):
        with Timer("Get Geo Anomalies Case Type: Parse Params"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []
            
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)

            yesterday = end_date - timedelta(days=1)
            thirty_days_ago = end_date - timedelta(days=30)
        with Timer("Get Geo Anomalies Case Type: Build Base queryset"):
            # Base queryset - only private hospitals with state mismatches
            anomalies = Last24Hour.objects.filter(
                hospital_type='P',
                patient_state_name__isnull=False,
                hosp_state_name__isnull=False
            ).exclude(patient_state_name=F('hosp_state_name'))
        with Timer("Get Geo Anomalies Case Type: Filter on Districts"):
            # Apply district filter (patient's district)
            if districts:
                anomalies = anomalies.filter(patient_district_name__in=districts)
        with Timer("Get Geo Anomalies Case Type: Today's count"):
            # Today's count
            today_anomalies = anomalies.filter(
                Q(admission_dt__date__gte=start_date) &
                Q(admission_dt__date__lte=end_date) | 
                Q(admission_dt__isnull=True, preauth_init_date__date__gte=start_date) &
                Q(admission_dt__isnull=True, preauth_init_date__date__lte=end_date)
            )
        with Timer("Get Geo Anomalies Case Type: Yesterday's Count"):
            # Yesterday's count
            yesterday_anomalies = anomalies.filter(
                Q(admission_dt__date=yesterday) | 
                Q(admission_dt__isnull=True, preauth_init_date__date=yesterday)
            )
        with Timer("Get Geo Anomalies Case Type: Last 30 days"):
            # Last 30 days count
            last_30_days_anomalies = anomalies.filter(
                Q(admission_dt__gte=thirty_days_ago) | 
                Q(admission_dt__isnull=True, preauth_init_date__gte=thirty_days_ago)
            )
        with Timer("Get Geo Anomalies Case Type: Filter State Mismatch Anomalies"):
            # Get state mismatch statistics
            state_mismatches = (
                anomalies.values('patient_state_name', 'hosp_state_name')
                .annotate(count=Count('id'))
                .order_by('-count')[:10]  # Top 10 mismatch pairs
            )

        data = {
            'total': today_anomalies.count(),
            'yesterday': yesterday_anomalies.count(),
            'last_30_days': last_30_days_anomalies.count(),
            'state_mismatches': list(state_mismatches)
        }

    return JsonResponse(data)

def get_geo_anomalies_details(request):
    with Timer("get_geo_anomalies_details TOTAL"):
        with Timer("Parse Params"):
            district_param = request.GET.get('district', '')
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 50))
            districts = district_param.split(',') if district_param else []
            
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
        with Timer("get_geo_anomalies_details: Build Base queryset"):
            cases = Last24Hour.objects.filter(
                hospital_type='P',
                patient_state_name__isnull=False,
                hosp_state_name__isnull=False
            ).exclude(patient_state_name=F('hosp_state_name')).filter(
                Q(admission_dt__date__gte=start_date) &
                Q(admission_dt__date__lte=end_date) | 
                Q(admission_dt__isnull=True, preauth_init_date__date__gte=start_date) &
                Q(admission_dt__isnull=True, preauth_init_date__date__lte=end_date)
            ).order_by('patient_state_name', 'hosp_state_name')
        with Timer("Filter on districts"):
            if districts:
                cases = cases.filter(patient_district_name__in=districts)
        with Timer("get_geo_anomalies_details: Pagination"):
            paginator = Paginator(cases, page_size)
            page_obj = paginator.get_page(page)
        with Timer("Serialize"):
            data = []
            for idx, case in enumerate(page_obj.object_list, 1):
                data.append({
                    'serial_no': (page_obj.number - 1) * page_size + idx,
                    'claim_id': case.registration_id or case.case_id or 'N/A',
                    'patient_name': case.patient_name or f"Patient {case.member_id}",
                    'district_name': case.patient_district_name or 'N/A',
                    'preauth_initiated_date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
                    'preauth_initiated_time': case.preauth_init_date.strftime('%H:%M:%S') if case.preauth_init_date else 'N/A',
                    'hospital_id': case.hospital_id or 'N/A',
                    'hospital_name': case.hospital_name or 'N/A',
                    'patient_state': case.patient_state_name or 'N/A',
                    'hospital_state': case.hosp_state_name or 'N/A',
                })

    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })

def get_geo_violations_by_state(request):
    with Timer("get_geo_violations_by_state TOTAL"):
        with Timer("get_geo_violations_by_state: Parse params"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []
            
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
        with Timer("get_geo_violations_by_state: Build Base queryset"):
            cases = (
                Last24Hour.objects
                .filter(
                    hospital_type='P',
                    patient_state_name__isnull=False,
                    hosp_state_name__isnull=False
                )
                .exclude(patient_state_name=F('hosp_state_name'))
                .filter(
                    (
                        Q(admission_dt__date__gte=start_date) &
                        Q(admission_dt__date__lte=end_date)
                    ) |
                    (
                        Q(admission_dt__isnull=True) &
                        Q(preauth_init_date__date__gte=start_date) &
                        Q(preauth_init_date__date__lte=end_date)
                    )
                )
            )
        with Timer("get_geo_violations_by_state: Filter on districts"):
            if districts:
                cases = cases.filter(patient_district_name__in=districts)
        with Timer("get_geo_violations_by_state: Group by state and hospital_state"):
            # Group by BOTH patient_state and hospital_state
            result = (
                cases.values('patient_state_name', 'hosp_state_name')
                .annotate(count=Count('id'))
                .order_by('-count')
            )
        with Timer("get_geo_violations_by_state: Build Lables"):
            # Build label like "Jharkhand → Bihar"
            states_pairs = [
                f"{item['patient_state_name']} → {item['hosp_state_name']}"
                for item in result
            ]
            counts = [item['count'] for item in result]
    
    return JsonResponse({
        'pairs': states_pairs,
        'counts': counts
    })

def get_geo_violations_demographics(request, type):
    with Timer("get_geo_violations_demographics TOTAL"):
        with Timer("get_geo_violations_demographics: Parse Params"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []

            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
        with Timer("get_geo_violations_demographics: Build base queryset"):
            base_query = Last24Hour.objects.filter(
                hospital_type='P',
                patient_state_name__isnull=False,
                hosp_state_name__isnull=False
                ).exclude(
                    patient_state_name=F('hosp_state_name')
                ).filter(
                    (
                        Q(admission_dt__date__gte=start_date) &
                        Q(admission_dt__date__lte=end_date)
                    ) |
                    (
                        Q(admission_dt__isnull=True) &
                        Q(preauth_init_date__date__gte=start_date) &
                        Q(preauth_init_date__date__lte=end_date)
                    )
                ).order_by('patient_state_name', 'hosp_state_name')
        with Timer("get_geo_violations_demographics: Filter on districts"):
            if districts:
                base_query = base_query.filter(patient_district_name__in=districts)
        with Timer("get_geo_violations_demographics: Filter on age and gender"):
            if type == 'age':
                with Timer("get_geo_violations_demographics: Group by age"):
                    age_groups = Case(
                        When(age__lt=20, then=Value('≤20')),
                        When(age__gte=20, age__lt=30, then=Value('21-30')),
                        When(age__gte=30, age__lt=40, then=Value('31-40')),
                        When(age__gte=40, age__lt=50, then=Value('41-50')),
                        When(age__gte=50, age__lt=60, then=Value('51-60')),
                        When(age__gte=60, then=Value('60+')),
                        default=Value('Unknown'),
                        output_field=CharField()
                    )
                    
                    age_data = base_query.annotate(age_group=age_groups).values('age_group') \
                        .annotate(count=Count('id')).order_by('age_group')
                    
                    categories = ['≤20', '21-30', '31-40', '41-50', '51-60', '60+', 'Unknown']
                    colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40', '#C9CBCF']
                    age_dict = {item['age_group']: item['count'] for item in age_data}
                    
                    data = {
                        'labels': categories,
                        'data': [age_dict.get(cat, 0) for cat in categories],
                        'colors': colors
                    }
                
            elif type == 'gender':
                with Timer("get_geo_violations_demographics: Group by Gender"):
                    gender_data = base_query.values('gender') \
                        .annotate(count=Count('id')).order_by('gender')
                    
                    categories = ['Male', 'Female', 'Other', 'Unknown']
                    colors = ['#36A2EB', '#FF6384', '#4BC0C0', '#C9CBCF']
                    gender_map = {'M': 'Male', 'F': 'Female', 'O': 'Other'}
                    gender_dict = {gender_map.get(item['gender'], 'Unknown'): item['count'] for item in gender_data}
                    
                    data = {
                        'labels': categories,
                        'data': [gender_dict.get(cat, 0) for cat in categories],
                        'colors': colors
                    }
    
    return JsonResponse(data)

def get_geo_violations_geo(request):
    with Timer("get_geo_violations_geo TOTAL"):
        with Timer("get_geo_violations_geo: Parse params"):
            district_param = request.GET.get('district', '')
            districts = district_param.split(',') if district_param else []

            # parse dates
            startDate = request.GET.get('start_date')
            endDate = request.GET.get('end_date')
            start_date, end_date = parse_date(startDate, endDate)
        with Timer("Build base queryset"):
            # base queryset
            qs = Last24Hour.objects.filter(
                hospital_type='P',
                patient_state_name__isnull=False,
                hosp_state_name__isnull=False
                ).exclude(
                    patient_state_name=F('hosp_state_name')
                ).filter(
                    (
                        Q(admission_dt__date__gte=start_date) &
                        Q(admission_dt__date__lte=end_date)
                    ) |
                    (
                        Q(admission_dt__isnull=True) &
                        Q(preauth_init_date__date__gte=start_date) &
                        Q(preauth_init_date__date__lte=end_date)
                    )
                )
        with Timer("get_geo_violations_geo: Filter on districts"):
            if districts:
                qs = qs.filter(patient_district_name__in=districts)
        with Timer("get_geo_violations_geo: Aggregate by id"):
            # aggregate by state_name
            agg = qs.values('patient_district_name').annotate(count=Count('id', distinct=True))
        with Timer("get_geo_violations_geo: Map to FID"):
            # map to FID
            result = []
            for row in agg:
                fid = SHAPEFILE_DISTRICT_MAPPING.get(row['patient_district_name'].lower())
                if fid is not None:
                    result.append({'fid': fid, 'count': row['count']})

    return JsonResponse(result, safe=False)

# Module-level cache
df_cache = None
capacity_map = None
CACHE_TTL = datetime.timedelta(minutes=5)





    

@require_http_methods(["GET"])
def download_high_value_claims_excel(request):
    # 1) read district filter
    district_param = request.GET.get('district', '')
    districts     = district_param.split(',') if district_param else []
   
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # 2) base queryset for P-type hospitals
    qs = high_value_claims_base_query(start_date, end_date)
    if districts:
        qs = qs.filter(patient_district_name__in=districts)

    # 3) split into surgical & medical
    surgical_qs = qs.filter(
        case_type__iexact='SURGICAL',
        amount_claim_initiated__gte=100000
    )
    medical_qs = qs.filter(
        case_type__iexact='MEDICAL',
        amount_claim_initiated__gte=25000
    )

    # 4) helper to serialize
    def serialize(qs):
        rows = []
        for idx, c in enumerate(qs, start=1):
            rows.append({
                'S.No':          idx,
                'Claim ID':      c.registration_id or c.case_id or 'N/A',
                'Patient Name':  c.patient_name or f"Patient {c.member_id}",
                'District':      c.patient_district_name or 'N/A',
                'Preauth Initiated Date': c.preauth_init_date.strftime('%Y-%m-%d') if c.preauth_init_date else 'N/A',
                'Preauth Initiated Time': c.preauth_init_date.strftime('%H:%M:%S') if c.preauth_init_date else 'N/A',
                'Hospital ID': c.hospital_id or 'N/A',
                'Hospital Name': c.hospital_name or 'N/A',
                'Amount':        c.amount_claim_initiated or 0,
                'Case Type':     c.case_type.upper() if c.case_type else 'N/A'
            })
        return rows

    df_surgical = pd.DataFrame(serialize(surgical_qs))
    df_medical  = pd.DataFrame(serialize(medical_qs))

    # print(df_medical)
    # 5) write to Excel with two sheets
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_surgical.to_excel(writer, index=False, sheet_name='Surgical Reports')
        df_medical.to_excel(writer, index=False, sheet_name='Medical Reports')

        wb   = writer.book
        ws_s = writer.sheets['Surgical Reports']
        ws_m = writer.sheets['Medical Reports']

        # 6) define styles
        red_fill   = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        blue_fill  = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
        white_font = Font(color='FFFFFFFF')

        # 7) style only the “Case Type” cells
        def style_sheet(ws, fill):
            # find the column index of “Case Type” in the header row
            header = next(ws.iter_rows(min_row=1, max_row=1))
            col_idx = next((i+1 for i, cell in enumerate(header) if cell.value == 'Case Type'), None)
            if not col_idx:
                return
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[col_idx-1]
                cell.fill = fill
                cell.font = white_font

        style_sheet(ws_s, red_fill)
        style_sheet(ws_m, blue_fill)

    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="high_value_claims_excel_report.xlsx"'
    return resp

@require_POST
@csrf_protect
def download_high_value_claims_report(request):
    # 1) Read inputs
    case_type      = request.POST.get('case_type', 'all').lower()   # 'all','surgical','medical'
    district_param = request.POST.get('district', '')
    districts      = [d for d in district_param.split(',') if d]
    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # 2) Helper for charts
    def strip_b64(key):
        val = request.POST.get(key, '')
        return val.split('base64,',1)[1] if 'base64,' in val else ''

    surgical_chart_b64     = strip_b64('surgical_chart')
    surgical_age_chart_b64 = strip_b64('surgical_age_chart')
    surgical_gen_chart_b64 = strip_b64('surgical_gender_chart')
    medical_chart_b64      = strip_b64('medical_chart')
    medical_age_chart_b64  = strip_b64('medical_age_chart')
    medical_gen_chart_b64  = strip_b64('medical_gender_chart')
    map_all_b64  = strip_b64('map_all')
    map_med_b64  = strip_b64('map_med')
    map_surg_b64 = strip_b64('map_surg')

    # Callouts
    surgical_age_callouts = request.POST.get('surgical_age_callouts','')
    surgical_gen_callouts = request.POST.get('surgical_gender_callouts','')
    medical_age_callouts  = request.POST.get('medical_age_callouts','')
    medical_gen_callouts  = request.POST.get('medical_gender_callouts','')

    # 3) Build querysets based on case_type
    base_qs = high_value_claims_base_query(start_date, end_date)
    surgical_qs = base_qs.none()
    medical_qs  = base_qs.none()

    if case_type in ['all','surgical']:
        surgical_qs = base_qs.filter(case_type__iexact='SURGICAL', amount_claim_initiated__gte=100000)
    if case_type in ['all','medical']:
        medical_qs  = base_qs.filter(case_type__iexact='MEDICAL',  amount_claim_initiated__gte=25000)

    if districts:
        surgical_qs = surgical_qs.filter(patient_district_name__in=districts)
        medical_qs  = medical_qs.filter(patient_district_name__in=districts)

    # 4) Serialize rows
    surgical_rows = [
        {
            'serial_no':     i+1,
            'claim_id':      c.registration_id or c.case_id or 'N/A',
            'patient_name':  c.patient_name or f"Patient {c.member_id}",
            'patient_district_name': c.patient_district_name or 'N/A',
            'preauth_initiated_date': c.preauth_init_date.strftime('%Y-%m-%d') if c.preauth_init_date else 'N/A',
            'preauth_initiated_time': c.preauth_init_date.strftime('%H:%M:%S') if c.preauth_init_date else 'N/A',
            'hospital_id': c.hospital_id or 'N/A',
            'hospital_name': c.hospital_name or 'N/A',
            'amount':        c.amount_claim_initiated or 0,
            'case_type':     'SURGICAL'
        }
        for i, c in enumerate(surgical_qs)
    ]
    medical_rows = [
        {
            'serial_no':     i+1,
            'claim_id':      c.registration_id or c.case_id or 'N/A',
            'patient_name':  c.patient_name or f"Patient {c.member_id}",
            'patient_district_name': c.patient_district_name or 'N/A',
            'preauth_initiated_date': c.preauth_init_date.strftime('%Y-%m-%d') if c.preauth_init_date else 'N/A',
            'preauth_initiated_time': c.preauth_init_date.strftime('%H:%M:%S') if c.preauth_init_date else 'N/A',
            'hospital_id': c.hospital_id or 'N/A',
            'hospital_name': c.hospital_name or 'N/A',
            'patient_district_name': c.patient_district_name or 'N/A',
            'amount':        c.amount_claim_initiated or 0,
            'case_type':     'MEDICAL'
        }
        for i, c in enumerate(medical_qs)
    ]

    # 5) Compute report_districts as a sorted list
    combined = [r['patient_district_name'] for r in surgical_rows + medical_rows if r['patient_district_name'] and r['patient_district_name'] != 'N/A']
    report_districts = sorted(set(combined))

    # 6) Render
    context = {
        'logo_url':                request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'title':                   'SAFU DASHBOARD ANALYSIS REPORT',
        'date':                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'case_type':               case_type,
        'report_districts':        report_districts,
        'surgical_rows':           surgical_rows,
        'medical_rows':            medical_rows,
        'surgical_chart_b64':      surgical_chart_b64,
        'surgical_age_chart_b64':  surgical_age_chart_b64,
        'surgical_gen_chart_b64':  surgical_gen_chart_b64,
        'medical_chart_b64':       medical_chart_b64,
        'medical_age_chart_b64':   medical_age_chart_b64,
        'medical_gen_chart_b64':   medical_gen_chart_b64,
        'surgical_age_callouts':   surgical_age_callouts,
        'surgical_gen_callouts':   surgical_gen_callouts,
        'medical_age_callouts':    medical_age_callouts,
        'medical_gen_callouts':    medical_gen_callouts,
        'map_all_b64':             map_all_b64,
        'map_med_b64':             map_med_b64,
        'map_surg_b64':            map_surg_b64,
    }
    html_string = render_to_string('high_value_claims_report.html', context)
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="High_Value_Claims_PDF_Report.pdf"'
    return response

@require_http_methods(["GET"])
def download_hospital_bed_cases_excel(request):
    # 1) District filter
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # 2) Date range
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # 3) Build bed_strength lookup
    bed_strengths = hospital_bed_cases_base_query()
    hospital_ids = list(bed_strengths.keys())

    # 4) Query admissions per hospital per day
    qs = (
        Last24Hour.objects
        .filter(
            hospital_type='P',
            hospital_id__in=hospital_ids,
            admission_dt__date__gte=start_date,
            admission_dt__date__lte=end_date,
        )
    )
    if districts:
        qs = qs.filter(patient_district_name__in=districts)

    admissions = (
        qs
        .annotate(day=TruncDate('admission_dt'))
        .values('hospital_id', 'hospital_name', 'patient_district_name', 'hosp_state_name', 'day')
        .annotate(admissions=Count('id'))
    )

    # 5) Flag violations (admissions > capacity on that day)
    violations = []
    for adm in admissions:
        cap = bed_strengths.get(adm['hospital_id'], 0)
        if adm['admissions'] > cap:
            violations.append({
                'S.No': len(violations) + 1,
                'Hospital ID': adm['hospital_id'],
                'Hospital Name': adm['hospital_name'],
                'District': adm['patient_district_name'],
                'State': adm['hosp_state_name'],
                'Bed Capacity': cap,
                'Admissions': adm['admissions'],
                'Excess': adm['admissions'] - cap,
                'Last Violation': adm['day'].strftime('%Y-%m-%d') if adm['day'] else 'N/A'
            })

    # 6) Build DataFrame
    df = pd.DataFrame(violations)

    # 7) Write to Excel with yellow fill on “Excess”
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hospital Bed Cases')
        ws = writer.sheets['Hospital Bed Cases']

        yellow = PatternFill(start_color='FFFF00',
                             end_color='FFFF00',
                             fill_type='solid')

        # find the column index for “Excess” in row 1
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if 'Excess' in headers:
            col_idx = headers.index('Excess') + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                row[0].fill = yellow

    buffer.seek(0)
    resp = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="hospital_bed_cases.xlsx"'
    return resp

@require_POST
@csrf_protect
def download_hospital_bed_report(request):
    # 1) Read inputs
    district_param = request.POST.get('district', '')
    districts = [d for d in district_param.split(',') if d]
    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # strip base64
    hc = request.POST.get('hospital_chart', '')
    hospital_chart_b64 = hc.split('base64,', 1)[1] if 'base64,' in hc else ''

    hc_map = request.POST.get('hospital_beds', '')
    hosp_bed_map_b64 = hc_map.split('base64,', 1)[1] if 'base64,' in hc_map else ''

    # 2) Build full violations list (no pagination)
    bed_strengths = hospital_bed_cases_base_query()
    hospital_ids = list(bed_strengths.keys())

    qs = (
        Last24Hour.objects
        .filter(
            hospital_type='P',
            hospital_id__in=hospital_ids,
            admission_dt__date__gte=start_date,
            admission_dt__date__lte=end_date,
        )
    )
    if districts:
        qs = qs.filter(patient_district_name__in=districts)

    admissions = (
        qs
        .annotate(day=TruncDate('admission_dt'))
        .values('hospital_id', 'hospital_name', 'patient_district_name', 'hosp_state_name', 'day')
        .annotate(admissions=Count('id'))
    )

    rows = []
    for i, adm in enumerate(admissions, start=1):
        cap = bed_strengths.get(adm['hospital_id'], 0)
        if adm['admissions'] > cap:
            rows.append({
                'serial_no':      i,
                'hospital_id':    adm['hospital_id'],
                'hospital_name':  adm['hospital_name'],
                'district':       adm['patient_district_name'],
                'state':          adm['hosp_state_name'],
                'bed_capacity':   cap,
                'admissions':     adm['admissions'],
                'excess':         adm['admissions'] - cap,
                'last_violation': adm['day'].strftime('%Y-%m-%d') if adm['day'] else 'N/A'
            })

    # 3) report_districts from those rows
    report_districts = sorted({r['district'] for r in rows})

    # 4) Render PDF
    context = {
        'logo_url':          request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'title':             'SAFU DASHBOARD ANALYSIS REPORT',
        'date':              datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'report_districts':  report_districts,
        'table_rows':        rows,
        'hospital_chart_b64': hospital_chart_b64,
        'hosp_bed_map_b64':   hosp_bed_map_b64,
    }
    html = render_to_string('hospital_bed_report.html', context)
    pdf  = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Hospital_Bed_Claims_PDF_Report.pdf"'
    return response

@require_http_methods(["GET"])
def download_family_id_cases_excel(request):
    # 1) district filter
    district_param = request.GET.get('district', '')
    districts     = district_param.split(',') if district_param else []

    # 2) date
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # 3) build the subquery & base queryset
    subq = (
        Last24Hour.objects
        .annotate(day=TruncDate('preauth_init_date'))
        .values('family_id', 'day')
        .annotate(count=Count('id'))
        .filter(count__gt=1)
        .values('family_id')
    )
    qs = (
        Last24Hour.objects
        .filter(
            Q(hospital_type='P'),
            Q(family_id__in=Subquery(subq)),
            Q(preauth_init_date__date__gte=start_date) &
            Q(preauth_init_date__date__lte=end_date)
        )
        .order_by('family_id', 'preauth_init_date')
    )
    if districts:
        qs = qs.filter(patient_district_name__in=districts)

    # 4) serialize all rows
    rows = []
    for idx, case in enumerate(qs, start=1):
        rows.append({
            'S.No':          idx,
            'Family ID':     case.family_id,
            'Claim ID':      case.registration_id or case.case_id or 'N/A',
            'Patient Name':  case.patient_name or f"Patient {case.member_id}",
            'District':      case.patient_district_name or 'N/A',
            'Preauth Initiated Date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
            'Preauth Initiated Time': case.preauth_init_date.strftime('%H:%M:%S') if case.preauth_init_date else 'N/A',
            'Hospital ID': case.hospital_id or 'N/A',
            'Hospital Name': case.hospital_name or 'N/A',
            'Date':          case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
        })

    # 5) build DataFrame
    df = pd.DataFrame(rows)

    # 6) write to Excel, then conditionally color “Family ID” cells
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Family ID Cases')
        ws = writer.sheets['Family ID Cases']

        # Only attempt coloring if DataFrame has that column
        if not df.empty and 'Family ID' in df.columns:
            # Map each family_id to a consistent random color
            family_ids = df['Family ID'].unique()
            color_map  = {}
            for fid in family_ids:
                random.seed(str(fid))
                r, g, b = (random.randint(0,255) for _ in range(3))
                color_map[fid] = "{:02X}{:02X}{:02X}".format(r, g, b)

            # Find the column index of “Family ID”
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if 'Family ID' in headers:
                col_idx = headers.index('Family ID') + 1
                yellow_fill = None  # placeholder

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell = row[col_idx-1]
                    fid  = cell.value
                    if fid in color_map:
                        fill = PatternFill(
                            start_color=color_map[fid],
                            end_color=color_map[fid],
                            fill_type='solid'
                        )
                        cell.fill = fill

    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="family_id_cases.xlsx"'
    return resp

@require_POST
@csrf_protect
def download_family_id_cases_report(request):
    # 1) inputs
    district_param = request.POST.get('district','')
    districts      = [d for d in district_param.split(',') if d]
    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # strip base64 helper
    def strip_b64(key):
        val = request.POST.get(key,'')
        return val.split('base64,',1)[1] if 'base64,' in val else ''

    family_chart_b64 = strip_b64('family_chart')
    age_b64          = strip_b64('family_age_chart')
    gender_b64       = strip_b64('family_gender_chart')
    age_callouts     = request.POST.get('age_callouts','')
    gender_callouts  = request.POST.get('gender_callouts','')
    family_id_map_b64 = strip_b64('family_id')

    # subquery families with >2 claims today
    freq_families = Last24Hour.objects.annotate(
        day=TruncDate('preauth_init_date')
    ).values('family_id','day') \
     .annotate(cnt=Count('id')).filter(cnt__gt=1) \
     .values('family_id')

    qs = Last24Hour.objects.filter(
        Q(hospital_type='P'),
        Q(family_id__in=Subquery(freq_families)),
        Q(preauth_init_date__date__gte=start_date) &
        Q(preauth_init_date__date__lte=end_date)
    ).order_by('family_id','preauth_init_date')

    if districts:
        qs = qs.filter(patient_district_name__in=districts)

    rows = []
    for idx, c in enumerate(qs, start=1):
        rows.append({
            'serial_no':    idx,
            'family_id':    c.family_id,
            'claim_id':     c.registration_id or c.case_id or 'N/A',
            'patient_name': c.patient_name or f"Patient {c.member_id}",
            'district':     c.patient_district_name or 'N/A',
            'preauth_initiated_date':     c.preauth_init_date.strftime('%Y-%m-%d') if c.preauth_init_date else 'N/A',
            'preauth_initiated_time':     c.preauth_init_date.strftime('%H:%M:%S') if c.preauth_init_date else 'N/A',
            'hospital_id':     c.hospital_id or 'N/A',
            'hospital_name':     c.hospital_name or 'N/A'
        })

    # 3) report_districts
    report_districts = sorted({r['district'] for r in rows if r['district'] != 'N/A'})

    # 4) render template
    context = {
        'logo_url':            request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'title':               'SAFU DASHBOARD ANALYSIS REPORT',
        'date':                datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'report_districts':    report_districts,
        'table_rows':          rows,
        'family_chart_b64':    family_chart_b64,
        'age_b64':             age_b64,
        'gender_b64':          gender_b64,
        'age_callouts':        age_callouts,
        'gender_callouts':     gender_callouts,
        'family_id_map_b64':   family_id_map_b64,
    }
    html_string = render_to_string('family_id_report.html', context)

    # 5) PDF
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="Family_ID_Cases_PDF_Report.pdf"'
    return resp

@require_http_methods(["GET"])
def download_geo_anomalies_excel(request):
    # 1) District filter
    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # 2) date
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # 3) Query exactly as in get_geo_anomalies_details, but no pagination
    qs = Last24Hour.objects.filter(
        hospital_type='P',
        patient_state_name__isnull=False,
        hosp_state_name__isnull=False
        ).exclude(
            patient_state_name=F('hosp_state_name')
        ).filter(
            (
                Q(admission_dt__date__gte=start_date) &
                Q(admission_dt__date__lte=end_date)
            ) |
            (
                Q(admission_dt__isnull=True) &
                Q(preauth_init_date__date__gte=start_date) &
                Q(preauth_init_date__date__lte=end_date)
            )
        ).order_by('patient_state_name', 'hosp_state_name')

    if districts:
        qs = qs.filter(patient_district_name__in=districts)

    # 4) Serialize all rows
    rows = []
    for idx, c in enumerate(qs, start=1):
        rows.append({
            'S.No':             idx,
            'Claim ID':         c.registration_id or c.case_id or 'N/A',
            'Patient Name':     c.patient_name or f"Patient {c.member_id}",
            'District':         c.patient_district_name or 'N/A',
            'Preauth Initiated Date':         c.preauth_init_date.strftime('%Y-%m-%d') if c.preauth_init_date else 'N/A',
            'Preauth Initiated Time':         c.preauth_init_date.strftime('%H:%M:%S') if c.preauth_init_date else 'N/A',
            'Hospital ID':    c.hospital_id or 'N/A',
            'Hospital Name':    c.hospital_name or 'N/A',
            'Patient State':    c.patient_state_name or 'N/A',
            'Hospital State':   c.hosp_state_name or 'N/A',
        })

    df = pd.DataFrame(rows)

    # 5) Write to Excel, coloring the two state columns per row
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Geo Anomalies')
        ws = writer.sheets['Geo Anomalies']

        # Locate column indexes
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            pat_idx = headers.index('Patient State') + 1
            hos_idx = headers.index('Hospital State') + 1
        except ValueError:
            pat_idx = hos_idx = None

        if pat_idx and hos_idx:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # generate a random color per row
                r, g, b = [random.randint(0,255) for _ in range(3)]
                hexcol = f"{r:02X}{g:02X}{b:02X}"
                fill = PatternFill(start_color=hexcol,
                                   end_color=hexcol,
                                   fill_type='solid')
                # apply to both cells in this row
                row[pat_idx-1].fill = fill
                row[hos_idx-1].fill = fill

    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="geographic_anomalies_excel.xlsx"'
    return resp

@require_POST
@csrf_protect
def download_geo_anomalies_pdf_report(request):
    # 1) Read inputs
    district_param = request.POST.get('district','')
    districts      = [d for d in district_param.split(',') if d]

    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # helper to strip base64
    def strip_b64(key):
        v = request.POST.get(key,'')
        return v.split('base64,',1)[1] if 'base64,' in v else ''

    geo_b64    = strip_b64('geo_chart')
    age_b64    = strip_b64('geo_age_chart')
    gender_b64 = strip_b64('geo_gender_chart')
    age_c      = request.POST.get('geo_age_callouts','')
    gen_c      = request.POST.get('geo_gender_callouts','')
    geo_map_b64 = strip_b64('geo_anomalies')

    # 2) Fetch full anomalies (no pagination)
    qs = Last24Hour.objects.filter(
        hospital_type='P',
        patient_state_name__isnull=False,
        hosp_state_name__isnull=False
        ).exclude(
            patient_state_name=F('hosp_state_name')
        ).filter(
            (
                Q(admission_dt__date__gte=start_date) &
                Q(admission_dt__date__lte=end_date)
            ) |
            (
                Q(admission_dt__isnull=True) &
                Q(preauth_init_date__date__gte=start_date) &
                Q(preauth_init_date__date__lte=end_date)
            )
        ).order_by('patient_state_name', 'hosp_state_name')

    # 3) Build rows
    rows = []
    for i, c in enumerate(qs, start=1):
        rows.append({
            'serial_no':    i,
            'claim_id':     c.registration_id or c.case_id or 'N/A',
            'patient_name': c.patient_name or f"Patient {c.member_id}",
            'district':     c.patient_district_name or 'N/A',
            'preauth_initiated_date':     c.preauth_init_date.strftime('%Y-%m-%d') if c.preauth_init_date else 'N/A',
            'preauth_initiated_time':     c.preauth_init_date.strftime('%H:%M:%S') if c.preauth_init_date else 'N/A',
            'hospital_id':c.hospital_id or 'N/A',
            'hospital_name':c.hospital_name or 'N/A',
            'patient_state':c.patient_state_name or 'N/A',
            'hospital_state':c.hosp_state_name or 'N/A',
        })

    # 4) District line
    report_districts = sorted({r['district'] for r in rows if r['district']!='N/A'})

    # 5) Render PDF
    context = {
      'logo_url':         request.build_absolute_uri('/static/images/pmjaylogo.png'),
      'title':            'SAFU DASHBOARD ANALYSIS REPORT',
      'date':             datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
      'report_districts': report_districts,
      'table_rows':       rows,
      'geo_chart_b64':    geo_b64,
      'age_chart_b64':    age_b64,
      'gender_chart_b64': gender_b64,
      'age_callouts':     age_c,
      'gender_callouts':  gen_c,
      'geo_map_b64':      geo_map_b64
    }
    html_string = render_to_string('geo_anomalies_report.html', context)
    pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="Geographic_Anomalies_PDF_Report.pdf"'
    return resp

@require_http_methods(["GET"])
def download_ophthalmology_excel(request):
    # import pandas as pd
    pass
    import json
    import io
    from django.http import HttpResponse
    from openpyxl.styles import PatternFill, Font

    # ---------------------------------------------------------
    # 1. Parse Parameters
    # ---------------------------------------------------------
    violation_type = request.GET.get('type', 'all')
    district_param = request.GET.get('district', '').strip()
    districts = [d.strip() for d in district_param.split(',')] if district_param else []
    
    # Parse column filters (optional, kept from your code)
    filters_json = request.GET.get('filters', '{}')
    try:
        column_filters = json.loads(filters_json) if filters_json and filters_json != '{}' else {}
    except json.JSONDecodeError:
        column_filters = {}

    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    # Assuming parse_date is a helper function you have defined elsewhere
    start_date, end_date = parse_date(startDate, endDate)

    # ---------------------------------------------------------
    # 2. Load Data & Filter
    # ---------------------------------------------------------
    # Assuming load_dataframes is defined elsewhere
    df, cap_map = load_dataframes()

    # Apply Base Filters
    mask = (
        df['hospital_type'].eq('P') &
        df['category_details'].str.contains('Opthalmology', case=False, na=False) &
        df['date'].between(start_date, end_date)
    )
    if districts:
        mask &= df['patient_district_name'].isin(districts)

    df_base = df.loc[mask].copy()

    # ---------------------------------------------------------
    # 3. Precompute Violation Flags (MATCHING DASHBOARD LOGIC)
    # ---------------------------------------------------------
    df_base['is_age'] = (df_base['age'] < 40).astype(int)
    # Corrected Time Logic: Before 8 AM or After 6 PM (18:00)
    df_base['is_time'] = ((df_base['preauth_hour'] < 8) | (df_base['preauth_hour'] >= 18)).astype(int)
    df_base['is_ot'] = df_base['ot_violation'].astype(int)
    
    # "is_any" flag for filtering rows that have at least one violation
    df_base['is_any'] = ((df_base['is_age'] + df_base['is_time'] + df_base['is_ot']) > 0).astype(int)

    # ---------------------------------------------------------
    # 4. Generate "Hospital Summary" Sheet Data (THE FIX)
    # ---------------------------------------------------------
    # Filter only rows that have violations for the summary
    df_violators = df_base[df_base['is_any'] == 1].copy()

    # Aggregate: Calculate sum of each violation type per hospital
    # Aggregate Age and Time by Hospital
    hospital_summary = df_violators.groupby(['hospital_id', 'hospital_name']).agg(
        age_violations=('is_age', 'sum'),
        time_violations=('is_time', 'sum')
    ).reset_index()

    # Safely Calculate EXACT OT Excess per hospital for Excel
    df_ot = df_base[df_base['ot_violation'] == True]
    if not df_ot.empty:
        daily_ot = df_ot.groupby(['hospital_id', 'date']).size().reset_index(name='daily_count')
        daily_ot['capacity'] = daily_ot['hospital_id'].map(cap_map).fillna(float('inf'))
        daily_ot['excess'] = (daily_ot['daily_count'] - daily_ot['capacity']).clip(lower=0)
        hospital_ot_excess = daily_ot.groupby('hospital_id')['excess'].sum().reset_index()
        hospital_ot_excess.rename(columns={'excess': 'ot_violations'}, inplace=True)
    else:
        hospital_ot_excess = pd.DataFrame(columns=['hospital_id', 'ot_violations'])

    # Merge OT Excess into summary
    hospital_summary = pd.merge(hospital_summary, hospital_ot_excess, on='hospital_id', how='left')
    hospital_summary['ot_violations'] = hospital_summary['ot_violations'].fillna(0)

    # CORRECT LOGIC: Total = Sum of the three parts (not count of rows)
    hospital_summary['total_violations'] = (
        hospital_summary['age_violations'] + 
        hospital_summary['time_violations'] + 
        hospital_summary['ot_violations']
    )

    # Calculate Share % based on Grand Total
    grand_total_violations = hospital_summary['total_violations'].sum()

    if grand_total_violations > 0:
        hospital_summary['share_percent'] = (
            (hospital_summary['total_violations'] / grand_total_violations) * 100
        ).round(2)
    else:
        hospital_summary['share_percent'] = 0.0

    # Sort by Total Violations descending
    hospital_summary = hospital_summary.sort_values(by='total_violations', ascending=False)

    # Rename columns to match the dashboard export format
    hospital_summary.rename(columns={
        'hospital_id': 'Hospital ID',
        'hospital_name': 'Hospital Name',
        'total_violations': 'Total Violations',
        'age_violations': 'Age Violations (<40)',
        'time_violations': 'Out of Hours (8pm-8am)',
        'ot_violations': 'OT Overload',
        'share_percent': '% Share'
    }, inplace=True)

    # ---------------------------------------------------------
    # 5. Prepare Detailed Data Sheets
    # ---------------------------------------------------------
    # We define masks to filter data for specific tabs
    violation_masks = {
        'age': df_base['is_age'] == 1,
        'preauth': df_base['is_time'] == 1,
        'ot': df_base['is_ot'] == 1,
        'multiple': (df_base['is_age'] + df_base['is_time'] + df_base['is_ot']) > 1,
        'all': df_base['is_any'] == 1
    }

    # Determine which sheets to generate
    if violation_type != 'all' and violation_type != 'hospital_summary':
        sheet_keys = [violation_type]
    else:
        # Default sheets for full report
        sheet_keys = ['all', 'age', 'ot', 'preauth', 'multiple']

    # Build hospital-district lookup from HospitalBeds for the Excel export.
    # hosp_district_name is not in the cached DataFrame, so we use the
    # HospitalBeds master table which is the authoritative source for district.
    excel_district_map = {
        rec['hospital_id']: rec['hospital_district'] or 'N/A'
        for rec in HospitalBeds.objects.values('hospital_id', 'hospital_district')
    }

    # Helper to create standard columns for detail sheets
    def prepare_detail_df(sub_df, sheet_type):
        base = pd.DataFrame({
            'S.No': range(1, len(sub_df)+1),
            'Claim ID': sub_df['registration_id'].fillna(sub_df['case_id']).fillna('N/A'),
            'Patient Name': sub_df['patient_name'].fillna(sub_df['member_id'].apply(lambda x: f"Patient {x}")),
            'Patient District': sub_df['patient_district_name'].fillna('N/A'),
            'Hospital ID': sub_df['hospital_id'].fillna('N/A'),
            'Hospital Name': sub_df['hospital_name'].fillna('N/A'),
            'Hospital District': sub_df['hospital_id'].map(excel_district_map).fillna('N/A'),
            'Amount': sub_df['amount_preauth_initiated'].fillna(0),
            'Preauth Time': sub_df['preauth_init_date'].dt.strftime("%H:%M:%S").fillna("N/A"),
            'Preauth Date': sub_df['preauth_init_date'].dt.strftime("%Y-%m-%d").fillna("N/A"),
        })
        
        # Add descriptive columns based on violations
        if sheet_type in ('all', 'age', 'multiple'):
            base['Age Violation'] = sub_df['age'].apply(lambda x: "Yes (<40)" if x < 40 else "No")
        
        if sheet_type in ('all', 'ot', 'multiple'):
             # Assuming 'ot_violation' is boolean or 0/1
            base['OT Violation'] = sub_df['ot_violation'].apply(lambda x: "Yes" if x else "No")
            
        if sheet_type in ('all', 'preauth', 'multiple'):
            base['Time Violation'] = sub_df['preauth_hour'].apply(lambda h: "Yes" if (h < 8 or h >= 18) else "No")

        return base

    # ---------------------------------------------------------
    # 6. Write to Excel
    # ---------------------------------------------------------
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # --- A. Write Summary Sheet ---
        # Only write summary if requested or if getting 'all'
        if violation_type in ['all', 'hospital_summary']:
            hospital_summary.to_excel(writer, index=False, sheet_name='Hospital Summary')
            ws = writer.sheets['Hospital Summary']
            
            # Simple Header Styling
            header_fill = PatternFill('solid', fgColor="DDDDDD")
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = max_length + 2

        # --- B. Write Detail Sheets ---
        # Only if we are not just requesting the summary
        if violation_type != 'hospital_summary':
            for key in sheet_keys:
                # Filter data for this sheet
                mask = violation_masks.get(key, df_base['is_any'] == 1)
                sub_df = df_base.loc[mask].copy()
                
                if sub_df.empty:
                    continue

                # Prepare readable columns
                final_df = prepare_detail_df(sub_df, key)

                # Write sheet
                sheet_title = key.replace('_', ' ').title()
                final_df.to_excel(writer, index=False, sheet_name=sheet_title)
                
                # Apply basic styling to details
                ws = writer.sheets[sheet_title]
                for cell in ws[1]:
                    cell.fill = PatternFill('solid', fgColor="DDDDDD")
                    cell.font = Font(bold=True)

    output.seek(0)
    filename = f"ophthalmology_report_{start_date}_to_{end_date}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@require_POST
@csrf_protect
def download_ophthalmology_pdf_report(request):
    # —1— Read filters
    violation_type = request.POST.get('violation_type', 'all')
    district_param = request.POST.get('district', '').strip()
    districts = [d for d in district_param.split(',') if d]

    # —2— Parse date range
    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    # —3— Helper to strip base64
    def strip_b64(key):
        v = request.POST.get(key, '')
        return v.split('base64,', 1)[1] if 'base64,' in v else ''

    # charts
    charts = {
        'age_chart':      strip_b64('age_chart'),
        'ot_chart':       strip_b64('ot_chart'),
        'preauth_chart':  strip_b64('preauth_chart'),
        'multiple_chart': strip_b64('multiple_chart'),
        'map_all':        strip_b64('map_all'),
        'map_age':        strip_b64('map_age'),
        'map_ot':         strip_b64('map_ot'),
        'map_preauth':    strip_b64('map_preauth'),
        'map_multiple':   strip_b64('map_multiple'),
    }

    pies = {}
    callouts = {}
    for sec in ['all','age','ot','preauth','multiple']:
        pies[f'{sec}_age']    = strip_b64(f'{sec}_age_chart')
        pies[f'{sec}_gender'] = strip_b64(f'{sec}_gender_chart')
        callouts[f'{sec}_age']    = request.POST.get(f'{sec}_age_callouts', '')
        callouts[f'{sec}_gender'] = request.POST.get(f'{sec}_gender_callouts', '')

    # —4— Base QS
    base_qs = Last24Hour.objects.filter(
        hospital_type='P',
        category_details__contains='Opthalmology',
        preauth_init_date__date__gte=start_date,
        preauth_init_date__date__lte=end_date
    )
    if districts:
        base_qs = base_qs.filter(patient_district_name__in=districts)

    # —5— Compute OT-overflow IDs (Option A: same logic everywhere)
    cap_map = {
        rec['hospital_id']: rec['number_of_surgeons'] * 30
        for rec in HospitalBeds.objects
                              .filter(number_of_surgeons__isnull=False)
                              .values('hospital_id','number_of_surgeons')
    }
    flagged_ot = set()
    for hid, cap in cap_map.items():
        qs_h = base_qs.filter(hospital_id=hid).order_by('preauth_init_date')
        total = qs_h.count()
        if total > cap:
            flagged_ot.update(qs_h.values_list('id', flat=True)[cap:])

    # —6— Section-specific QS with correct MULTIPLE logic
    def section_qs(vtype):
        m_age = base_qs.filter(age__lt=40)
        m_ot = base_qs.filter(id__in=flagged_ot)
        m_preauth = base_qs.exclude(
            preauth_init_date__hour__gte=8,
            preauth_init_date__hour__lt=18
        )

        if vtype == 'age':
            return m_age
        if vtype == 'ot':
            return m_ot
        if vtype == 'preauth':
            return m_preauth
        if vtype == 'multiple':
            # IDs that violate more than one rule
            ids = list(m_age.values_list('id', flat=True)) \
                + list(m_ot.values_list('id', flat=True)) \
                + list(m_preauth.values_list('id', flat=True))
            from collections import Counter
            c = Counter(ids)
            multi_ids = [cid for cid, cnt in c.items() if cnt > 1]
            return base_qs.filter(id__in=multi_ids)

        # 'all'
        return base_qs.filter(
            Q(age__lt=40) |
            Q(id__in=flagged_ot) |
            ~Q(preauth_init_date__hour__gte=8, preauth_init_date__hour__lt=18)
        )

    # —7— Build row tables
    rows = {}
    for sec in ['all','age','ot','preauth','multiple']:
        qs_iter = section_qs(sec).order_by('preauth_init_date')
        rows[sec] = [
            {
                'serial_no':    i + 1,
                'claim_id':     c.registration_id or c.case_id or 'N/A',
                'patient_name': c.patient_name or f"Patient {c.member_id}",
                'district':     c.patient_district_name or 'N/A',
                'hospital_id':  c.hospital_id or 'N/A',
                'hospital_name':c.hospital_name or 'N/A',
                'amount':       getattr(c,'amount_preauth_initiated',0) or 0,
                'age_lt_40':    bool(c.age and c.age < 40),
                'ot_cases':     c.id in flagged_ot,
                'preauth_time': bool(
                    c.preauth_init_date and
                    (c.preauth_init_date.hour < 8 or c.preauth_init_date.hour >= 18)
                ),
                'num_true': sum([
                    bool(c.age and c.age < 40),
                    c.id in flagged_ot,
                    bool(
                        c.preauth_init_date and
                        (c.preauth_init_date.hour < 8 or c.preauth_init_date.hour >= 18)
                    )
                ])
            }
            for i, c in enumerate(qs_iter)
        ]

    report_districts = sorted({
        r['district'] for r in rows['all'] if r['district'] != 'N/A'
    })

    # —8— Render HTML → PDF
    context = {
        'logo_url': request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'title': 'SAFU DASHBOARD ANALYSIS REPORT',
        'date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'violation_type': violation_type,
        'report_districts': report_districts,
        'start_date': start_date,
        'end_date': end_date,
        **{f'{sec}_rows': rows[sec] for sec in rows},
        **{f'{k}_b64': v for k, v in charts.items()},
        **{f'{k}_b64': v for k, v in pies.items()},
        **{f'{k}_callouts': v for k, v in callouts.items()},
    }

    html = render_to_string('ophthalmology_report.html', context)
    pdf = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()

    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = (
        f'attachment; filename="Ophthalmology_Report_{start_date}_to_{end_date}.pdf"'
    )
    return resp

def high_alert_total_count(request):
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    district_param = request.GET.get('district', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    districts = district_param.split(',') if district_param else []

    # Get watchlist hospitals
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)

    # Base query with same filters as get_flagged_claims_details
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_init_date__date__gte=start_date) & Q(preauth_init_date__date__lte=end_date) |
         Q(admission_dt__date__gte=start_date) & Q(admission_dt__date__lte=end_date))
    )

    if districts:
        base_query = base_query.filter(patient_district_name__in=districts)

    flagged_ot_ids = get_ot_overflow_hospital_ids(start_date, end_date, districts)
    annotated_cases = base_query.annotate(
        # Renamed annotations to avoid field conflicts
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', amount_claim_initiated__gte=100000) |
                Q(case_type__iexact='MEDICAL', amount_claim_initiated__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_dt__date__gte=start_date,
                                    admission_dt__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_init_date__date__gte=start_date,
                        preauth_init_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(patient_state_name=F('hosp_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ot_overflow=Case(
            When(id__in=flagged_ot_ids, then=Value(True)),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case = Case(
            When(
                Q(category_details__contains='Opthalmology') & (
                    Q(age__lt=40) |
                    Q(id__in=flagged_ot_ids) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__lt=8) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        )
    )

    # Calculate total flags
    filtered_cases = annotated_cases.annotate(
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    total = filtered_cases.count()

    return JsonResponse({"total": total})

def high_alert(request):
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    district_param = request.GET.get('district', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    districts = district_param.split(',') if district_param else []

    # Get watchlist hospitals
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)

    # Base query with same filters as get_flagged_claims_details
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_init_date__date__gte=start_date) & Q(preauth_init_date__date__lte=end_date) |
         Q(admission_dt__date__gte=start_date) & Q(admission_dt__date__lte=end_date))
    )

    if districts:
        base_query = base_query.filter(patient_district_name__in=districts)

    flagged_ot_ids = get_ot_overflow_hospital_ids(start_date, end_date, districts)
    annotated_cases = base_query.annotate(
        # Renamed annotations to avoid field conflicts
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', amount_claim_initiated__gte=100000) |
                Q(case_type__iexact='MEDICAL', amount_claim_initiated__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_dt__date__gte=start_date,
                                    admission_dt__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_init_date__date__gte=start_date,
                        preauth_init_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(patient_state_name=F('hosp_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ot_overflow=Case(
            When(id__in=flagged_ot_ids, then=Value(True)),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case = Case(
            When(
                Q(category_details__contains='Opthalmology') & (
                    Q(age__lt=40) |
                    Q(id__in=flagged_ot_ids) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__lt=8) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        )
    )

    # Calculate total flags
    filtered_cases = annotated_cases.annotate(
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    # Pagination
    paginator = Paginator(filtered_cases.order_by('-preauth_init_date'), page_size)
    page_obj = paginator.get_page(page)

    # Prepare response data
    data = []
    for idx, case in enumerate(page_obj.object_list, 1):
        data.append({
            'serial_no': (page_obj.number - 1) * page_size + idx,
            'claim_id': case.registration_id or case.case_id or 'N/A',
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'hospital_id': case.hospital_id,
            'hospital_name': case.hospital_name or 'N/A',
            'district': case.patient_district_name or 'N/A',
            'preauth_initiated_date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
            'preauth_initiated_time': case.preauth_init_date.strftime('%H:%M:%S') if case.preauth_init_date else 'N/A',
            'watchlist_hospital': '✓' if case.is_watchlist else '',
            'high_value_claims': '✓' if case.is_high_value else '',
            'hospital_bed_cases': '✓' if case.is_hospital_bed else '',
            'family_id_cases': '✓' if case.is_family_case else '',
            'geographic_anomalies': '✓' if case.is_geo_anomaly else '',
            'ophthalmology_cases': '✓' if case.is_ophtha_case else '',
        })

    return JsonResponse({
        'data': data,
        'pagination': {
            'total_records': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous()
        }
    })

def high_alert_district_distribution(request):
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # Reuse same filtering as high_alert view
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_init_date__date__gte=start_date) & Q(preauth_init_date__date__lte=end_date) | Q(admission_dt__date__gte=start_date) & Q(admission_dt__date__lte=end_date)))
    if districts:
        base_query = base_query.filter(patient_district_name__in=districts)

    flagged_ot_ids = get_ot_overflow_hospital_ids(start_date, end_date, districts)
    # Apply full high alert criteria annotations
    annotated = base_query.annotate(
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', amount_claim_initiated__gte=100000) |
                Q(case_type__iexact='MEDICAL', amount_claim_initiated__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_dt__date__gte=start_date,
                                    admission_dt__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_init_date__date__gte=start_date,
                        preauth_init_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(patient_state_name=F('hosp_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case = Case(
            When(
                Q(category_details__contains='Opthalmology') & (
                    Q(age__lt=40) |
                    Q(id__in=flagged_ot_ids) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__lt=8) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    # Aggregate by district
    result = annotated.values('patient_district_name').annotate(
        case_count=Count('id')
    ).order_by('-case_count')

    return JsonResponse({
        'labels': [d['patient_district_name'] or 'Unknown' for d in result],
        'counts': [d['case_count'] for d in result]
    })

def high_alert_demographics(request, type):
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # Reuse high alert base query
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_init_date__date__gte=start_date) & Q(preauth_init_date__date__lte=end_date) | Q(admission_dt__date__gte=start_date) & Q(admission_dt__date__lte=end_date))
    )
    
    if districts:
        base_query = base_query.filter(patient_district_name__in=districts)

    flagged_ot_ids = get_ot_overflow_hospital_ids(start_date, end_date, districts)
    # Apply high alert criteria annotations
    annotated = base_query.annotate(
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', amount_claim_initiated__gte=100000) |
                Q(case_type__iexact='MEDICAL', amount_claim_initiated__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_dt__date__gte=start_date,
                                    admission_dt__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_init_date__date__gte=start_date,
                        preauth_init_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(patient_state_name=F('hosp_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case = Case(
            When(
                Q(category_details__contains='Opthalmology') & (
                    Q(age__lt=40) |
                    Q(id__in=flagged_ot_ids) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__lt=8) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    if type == 'age':
        age_groups = Case(
            When(age__lt=20, then=Value('≤20')),
            When(age__range=(20, 29), then=Value('21-30')),
            When(age__range=(30, 39), then=Value('31-40')),
            When(age__range=(40, 49), then=Value('41-50')),
            When(age__range=(50, 59), then=Value('51-60')),
            When(age__gte=60, then=Value('60+')),
            default=Value('Unknown'),
            output_field=CharField()
        )
        
        data = annotated.annotate(age_group=age_groups).values('age_group') \
            .annotate(count=Count('id')).order_by('age_group')
        
        categories = ['≤20', '21-30', '31-40', '41-50', '51-60', '60+', 'Unknown']
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40', '#C9CBCF']
        
        result = {item['age_group']: item['count'] for item in data}
        return JsonResponse({
            'labels': categories,
            'data': [result.get(cat, 0) for cat in categories],
            'colors': colors
        })

    elif type == 'gender':
        gender_data = annotated.values('gender') \
            .annotate(count=Count('id')).order_by('gender')
        
        categories = ['Male', 'Female', 'Other', 'Unknown']
        colors = ['#36A2EB', '#FF6384', '#4BC0C0', '#C9CBCF']
        gender_map = {'M': 'Male', 'F': 'Female', 'O': 'Other'}
        
        result = defaultdict(int)
        for item in gender_data:
            gender = gender_map.get(item['gender'], 'Unknown')
            result[gender] += item['count']
        
        return JsonResponse({
            'labels': categories,
            'data': [result[cat] for cat in categories],
            'colors': colors
        })
    
def high_alerts_geo(request):
    """
    Returns district-wise high alert counts mapped to FID for ArcGIS map.
    """
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # Reuse same filtering as high_alert view
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (
            Q(preauth_init_date__date__gte=start_date) & Q(preauth_init_date__date__lte=end_date) |
            Q(admission_dt__date__gte=start_date) & Q(admission_dt__date__lte=end_date)
        )
    )
    if districts:
        base_query = base_query.filter(patient_district_name__in=districts)

    flagged_ot_ids = get_ot_overflow_hospital_ids(start_date, end_date, districts)
    # Apply full high alert criteria annotations
    annotated = base_query.annotate(
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', amount_claim_initiated__gte=100000) |
                Q(case_type__iexact='MEDICAL', amount_claim_initiated__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_dt__date__gte=start_date,
                                    admission_dt__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_init_date__date__gte=start_date,
                        preauth_init_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(patient_state_name=F('hosp_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case = Case(
            When(
                Q(category_details__contains='Opthalmology') & (
                    Q(age__lt=40) |
                    Q(id__in=flagged_ot_ids) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__lt=8) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    # Aggregate by district
    result = annotated.values('patient_district_name').annotate(
        case_count=Count('id')
    ).order_by('-case_count')

    geo_data = []
    for row in result:
        patient_district_name = row['patient_district_name']
        cnt = row['case_count']
        fid = SHAPEFILE_DISTRICT_MAPPING.get((patient_district_name or '').lower())
        if fid is not None:
            geo_data.append({'fid': fid, 'count': cnt})
    # After building geo_data
    for name, fid in SHAPEFILE_DISTRICT_MAPPING.items():
        if not any(d['fid'] == fid for d in geo_data):
            geo_data.append({'fid': fid, 'count': 0})

    return JsonResponse(geo_data, safe=False)

def download_high_alerts_excel(request):
    startDate = request.GET.get('start_date')
    endDate = request.GET.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)

    district_param = request.GET.get('district', '')
    districts = district_param.split(',') if district_param else []

    # Reuse same filtering logic as high_alert view
    suspicious_hospitals = SuspiciousHospital.objects.values_list('hospital_id', flat=True)
    
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=suspicious_hospitals) &
        Q(hospital_type='P') &
        (Q(preauth_init_date__date__gte=start_date) & Q(preauth_init_date__date__lte=end_date) | Q(admission_dt__date__gte=start_date) & Q(admission_dt__date__lte=end_date)))
    
    if districts:
        base_query = base_query.filter(patient_district_name__in=districts)

    flagged_ot_ids = get_ot_overflow_hospital_ids(start_date, end_date, districts)
    # Annotate cases with same logic as high_alert view
    annotated_cases = base_query.annotate(
        # Renamed annotations to avoid field conflicts
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', amount_claim_initiated__gte=100000) |
                Q(case_type__iexact='MEDICAL', amount_claim_initiated__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_dt__date__gte=start_date,
                                    admission_dt__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_init_date__date__gte=start_date,
                        preauth_init_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(patient_state_name=F('hosp_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case = Case(
            When(
                Q(category_details__contains='Opthalmology') & (
                    Q(age__lt=40) |
                    Q(id__in=flagged_ot_ids) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__lt=8) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        )
    )

    # Calculate total flags
    filtered_cases = annotated_cases.annotate(
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2)

    # Prepare data with new columns
    rows = []
    for case in annotated_cases.order_by('-preauth_init_date'):
        rows.append({
            'Serial No': '',  # Will be regenerated
            'Claim ID': case.registration_id or case.case_id or 'N/A',
            'Patient Name': case.patient_name or f"Patient {case.member_id}",
            'Hospital ID': case.hospital_id,
            'Hospital Name': case.hospital_name or 'N/A',
            'District': case.patient_district_name or 'N/A',
            'Preauth Initiated Date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
            'Preauth Initiated Time': case.preauth_init_date.strftime('%H:%M:%S') if case.preauth_init_date else 'N/A',
            'Watchlist': case.is_watchlist if case.is_watchlist else '',
            'High Value': case.is_high_value if case.is_high_value else '',
            'Bed Cases': case.is_hospital_bed if case.is_hospital_bed else '',
            'Family ID': case.is_family_case if case.is_family_case else '',
            'Geo Anomaly': case.is_geo_anomaly if case.is_geo_anomaly else '',
            'Ophthalmology': case.is_ophtha_case if case.is_ophtha_case else ''
        })

    # Create DataFrame with correct column order
    columns = [
        'Serial No', 'Claim ID', 'Patient Name', 'Hospital ID', 'Hospital Name',
        'District', 'Preauth Initiated Date', 'Preauth Initiated Time',
        'Watchlist', 'High Value', 'Bed Cases', 'Family ID', 'Geo Anomaly', 'Ophthalmology'
    ]
    df = pd.DataFrame(rows, columns=columns)
    df['Serial No'] = df.index + 1  # Continuous numbering

    # Excel styling
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='High Alerts')
        workbook = writer.book
        worksheet = writer.sheets['High Alerts']
        
        # Define color mappings (Excel color codes)
        colors = {
            'Watchlist': '26547D',     # Blue
            'High Value': 'ef436b',    # Red
            'Bed Cases': 'ffce5c',     # Yellow
            'Family ID': '05c793',     # Green
            'Geo Anomaly': '0091b9',   # Dark Blue
            'Ophthalmology': '1abc9c'  # Teal
        }

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply cell coloring
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in colors:
                fill = PatternFill(
                    start_color=colors[col_name],
                    end_color=colors[col_name],
                    fill_type='solid'
                )
                for row in worksheet.iter_rows(
                    min_row=2,
                    max_row=worksheet.max_row,
                    min_col=col_idx,
                    max_col=col_idx
                ):
                    for cell in row:
                        if cell.value:  # Only color if True
                            cell.fill = fill
                            cell.border = thin_border
                            cell.value = ''  # Optional: Clear boolean

    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"high_alerts_{start_date} to_{end_date}_{'_'.join(districts) if districts else 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

from django.views.decorators.csrf import ensure_csrf_cookie
@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def download_high_alert_report(request):
    # Get filters and chart data
    district = request.POST.get('district', '')
    districts = district.split(',') if district else []

    startDate = request.POST.get('start_date')
    endDate = request.POST.get('end_date')
    start_date, end_date = parse_date(startDate, endDate)
    
    # Process chart images
    def strip_prefix(data_url):
        return data_url.split('base64,', 1)[1] if data_url else ''
    
    district_b64 = strip_prefix(request.POST.get('district_chart', ''))
    age_b64 = strip_prefix(request.POST.get('age_chart', ''))
    gender_b64 = strip_prefix(request.POST.get('gender_chart', ''))
    map_b64 = strip_prefix(request.POST.get('map_image', ''))
    
    # Fetch all high alert cases
    base_query = Last24Hour.objects.filter(
        Q(hospital_id__in=SuspiciousHospital.objects.values_list('hospital_id', flat=True)) &
        Q(hospital_type='P') &
        (Q(preauth_init_date__date__gte=start_date) & Q(preauth_init_date__date__lte=end_date) | Q(admission_dt__date__gte=start_date) & Q(admission_dt__date__lte=end_date))
    )
    
    if districts:
        base_query = base_query.filter(patient_district_name__in=districts)
    
    flagged_ot_ids = get_ot_overflow_hospital_ids(start_date, end_date, districts)
    # Annotate and filter
    cases = base_query.annotate(
        # Renamed annotations to avoid field conflicts
        is_watchlist=Value(True, output_field=BooleanField()),
        is_high_value=Case(
            When(
                Q(case_type__iexact='SURGICAL', amount_claim_initiated__gte=100000) |
                Q(case_type__iexact='MEDICAL', amount_claim_initiated__gte=25000),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_hospital_bed=Case(
            When(
                Exists(
                    HospitalBeds.objects.filter(
                        hospital_id=OuterRef('hospital_id')
                    ).annotate(
                        admissions=Count(
                            'id',
                            filter=Exists(
                                Last24Hour.objects.filter(
                                    hospital_id=OuterRef('hospital_id'),
                                    admission_dt__date__gte=start_date,
                                    admission_dt__date__lte=end_date
                                )
                            )
                        )
                    ).filter(admissions__gt=F('bed_strength'))
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_family_case=Case(
            When(
                Exists(
                    Last24Hour.objects.filter(
                        family_id=OuterRef('family_id'),
                        preauth_init_date__date__gte=start_date,
                        preauth_init_date__date__lte=end_date
                    ).values('family_id').annotate(
                        count=Count('id')
                    ).filter(count__gt=1)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_geo_anomaly=Case(
            When(
                ~Q(patient_state_name=F('hosp_state_name')),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        ),
        is_ophtha_case = Case(
            When(
                Q(category_details__contains='Opthalmology') & (
                    Q(age__lt=40) |
                    Q(id__in=flagged_ot_ids) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__lt=8) |
                    Q(preauth_init_date__isnull=False, preauth_init_date__hour__gte=18)
                ),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        )
    )

    # Calculate total flags
    filtered_cases = cases.annotate(
        total_flags=(
            Cast('is_watchlist', IntegerField()) +
            Cast('is_high_value', IntegerField()) +
            Cast('is_hospital_bed', IntegerField()) +
            Cast('is_family_case', IntegerField()) +
            Cast('is_geo_anomaly', IntegerField()) +
            Cast('is_ophtha_case', IntegerField())
        )
    ).filter(total_flags__gte=2).order_by('-preauth_init_date')
    # Prepare table data
    table_rows = []
    for idx, case in enumerate(cases, 1):
        triggered_flags = []
        if case.is_watchlist: triggered_flags.append("Watchlist Hospitals")
        if case.is_high_value: triggered_flags.append("High Value Claims")
        if case.is_hospital_bed: triggered_flags.append("Hospital Bed Violations")
        if case.is_family_case: triggered_flags.append("Family ID Violations")
        if case.is_geo_anomaly: triggered_flags.append("Geo Anomaly")
        if case.is_ophtha_case: triggered_flags.append("Ophthalmology")
        
        table_rows.append({
            'serial_no': idx,
            'claim_id': case.registration_id or case.case_id,
            'patient_name': case.patient_name or f"Patient {case.member_id}",
            'district': case.patient_district_name,
            'hospital_id': case.hospital_id,
            'hospital_name': case.hospital_name,
            'triggered_flags': triggered_flags,
            'preauth_initiated_date': case.preauth_init_date.strftime('%Y-%m-%d') if case.preauth_init_date else 'N/A',
            'preauth_initiated_time': case.preauth_init_date.strftime('%H:%M:%S') if case.preauth_init_date else 'N/A',
        })
    
    # Render PDF
    context = {
        'logo_url': request.build_absolute_uri('/static/images/pmjaylogo.png'),
        'report_districts': list(set([case['district'] for case in table_rows if case['district']])),
        'table_rows': table_rows,
        'district_b64': district_b64,
        'age_b64': age_b64,
        'gender_b64': gender_b64,
        'map_b64': map_b64,
        'date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    html = render_to_string('high_alert_report.html', context)
    pdf = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="high_alert_report.pdf"'
    return response

@login_required(login_url='login')
def dashboard_view(request):
    # capture the district GET-param so your JS can pick it up
    district_param = request.GET.get('district', '')
    return render(request, 'dashboard.html', {
        'district_param': district_param,
        'active_page': 'dashboard',
    })

@login_required(login_url='login')
def high_alert_view(request):
    district_param = request.GET.get('district', '')
    return render(request, 'high_alert.html', {
        'district_param': district_param,
        'active_page': 'high_alert',
})
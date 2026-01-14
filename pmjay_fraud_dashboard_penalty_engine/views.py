"""
Penalty management views.
Validate input → call service → return JSON.
No business logic here.
"""

import json
import io
import logging
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from .services import (
    send_penalty_reminder,
    mark_penalty_paid,
    mark_penalty_non_compliant,
    close_penalty_case,
    get_penalties_page,
)
from .selectors import (
    get_penalty,
    list_penalty_audit_logs,
    get_penalty_summary,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Management page
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def management_page(request):
    """GET /penalty/management/  — Renders the penalty management page shell."""
    return render(request, 'penalty_case_maangement_template.html', {
        'active_page': 'penalty',
    })


# ─────────────────────────────────────────────────────────────────────────────
# List penalties API
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_GET
def list_penalties_view(request):
    """
    GET /api/penalty/cases/
    Query params:
        status       — ACTIVE | REMINDER_SENT | NON_COMPLIANT | PAID | CLOSED | ALL
        penalty_type — PENALTY | SUSPENSION | ALL
        search       — hospital_id or name substring
        page         — default 1
        page_size    — default 25, max 100
    """
    status       = request.GET.get('status', '').strip() or None
    penalty_type = request.GET.get('penalty_type', '').strip() or None
    search       = request.GET.get('search', '').strip() or None
    page_size    = min(int(request.GET.get('page_size', 25)), 100)
    page         = max(int(request.GET.get('page', 1)), 1)

    result = get_penalties_page(
        status=status,
        penalty_type=penalty_type,
        search=search,
        page=page,
        page_size=page_size,
    )
    return JsonResponse(result)


# ─────────────────────────────────────────────────────────────────────────────
# Summary counts
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_GET
def summary(request):
    """GET /api/penalty/summary/"""
    return JsonResponse(get_penalty_summary())


# ─────────────────────────────────────────────────────────────────────────────
# Audit log
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_GET
def penalty_audit_log(request, penalty_id):
    """GET /api/penalty/<id>/audit-log/"""
    penalty = get_penalty(penalty_id)
    if penalty is None:
        return JsonResponse({'error': f'Penalty {penalty_id} not found.'}, status=404)

    return JsonResponse({
        'penalty_id':    penalty_id,
        'hospital_name': penalty.hospital_name,
        'hospital_id':   penalty.hospital_id,
        'logs':          list_penalty_audit_logs(penalty_id),
    })


# ─────────────────────────────────────────────────────────────────────────────
# Lifecycle action endpoints
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def reminder(request, penalty_id):
    """POST /api/penalty/<id>/reminder/"""
    result = send_penalty_reminder(
        penalty_id=penalty_id,
        performed_by=request.user.username,
    )
    return JsonResponse(result, status=200 if result['ok'] else 400)


@login_required
@require_POST
def mark_paid(request, penalty_id):
    """POST /api/penalty/<id>/mark-paid/"""
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        data = {}

    result = mark_penalty_paid(
        penalty_id=penalty_id,
        performed_by=request.user.username,
        notes=data.get('notes', ''),
    )
    return JsonResponse(result, status=200 if result['ok'] else 400)


@login_required
@require_POST
def mark_non_compliant(request, penalty_id):
    """POST /api/penalty/<id>/non-compliant/"""
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        data = {}

    result = mark_penalty_non_compliant(
        penalty_id=penalty_id,
        performed_by=request.user.username,
        notes=data.get('notes', ''),
    )
    return JsonResponse(result, status=200 if result['ok'] else 400)


@login_required
@require_POST
def close(request, penalty_id):
    """POST /api/penalty/<id>/close/"""
    try:
        data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        data = {}

    result = close_penalty_case(
        penalty_id=penalty_id,
        performed_by=request.user.username,
        notes=data.get('notes', ''),
    )
    return JsonResponse(result, status=200 if result['ok'] else 400)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@require_GET
def export_excel(request):
    """GET /api/penalty/export/ — Export all penalty cases as XLSX."""
    import pandas as pd
    from openpyxl.styles import PatternFill, Font

    status       = request.GET.get('status', '').strip() or None
    penalty_type = request.GET.get('penalty_type', '').strip() or None
    search       = request.GET.get('search', '').strip() or None

    from .selectors import list_penalties, serialize_penalty
    qs = list_penalties(status=status, penalty_type=penalty_type, search=search)

    rows = []
    for p in qs:
        rows.append({
            'Case ID':           f"PEN-{p.pk}",
            'Hospital ID':       p.hospital_id,
            'Hospital Name':     p.hospital_name,
            'District':          p.district_name or 'N/A',
            'State':             p.state_name or 'N/A',
            'Type':              p.get_penalty_type_display(),  # type: ignore[attr-defined]
            'Status':            p.get_status_display(),  # type: ignore[attr-defined]
            'Penalty Amount (INR)': str(p.penalty_amount) if p.penalty_amount else 'N/A',
            'Suspension Period': p.suspension_label,
            'Imposed On':        p.imposed_at.strftime('%d %b %Y'),
            'Reminder Sent':     p.reminder_at.strftime('%d %b %Y') if p.reminder_at else '—',
            'Resolved On':       p.resolved_at.strftime('%d %b %Y') if p.resolved_at else '—',
            'Imposed By':        p.created_by,
            'Notes':             p.notes,
        })

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Penalty Cases')
        ws = writer.sheets['Penalty Cases']
        header_fill = PatternFill('solid', fgColor='DDDDDD')
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="penalty_cases.xlsx"'
    return response